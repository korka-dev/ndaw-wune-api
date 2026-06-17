"""Endpoints Admin — Élèves."""
from __future__ import annotations

import csv
import io
import uuid as uuid_module
from datetime import datetime, timezone
from typing import Optional

import pdfplumber
import openpyxl
from fastapi import APIRouter, Body, File, HTTPException, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, or_, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import selectinload

from app.core.deps import AdminUser, DB
from app.core.export_utils import build_xlsx_response
from app.core.pagination import Page, Pagination
from app.models.eleve import Eleve
from app.models.school import School
from app.schemas.eleve import EleveCreate, EleveUpdate, EleveResponse

router = APIRouter(prefix="/eleves", tags=["Admin — Élèves"])

# ── Colonnes attendues (acceptées en plusieurs langues / variantes) ────────────
# Clé = nom normalisé interne, valeurs = alias acceptés dans le fichier importé

COL_ALIASES: dict[str, list[str]] = {
    "nom":            ["nom", "name", "last_name", "lastname", "surname"],
    "prenom":         ["prenom", "prénom", "firstname", "first_name", "given_name"],
    "genre":          ["sexe", "genre", "sex", "gender"],
    "date_naissance": ["date_naissance", "naissance", "date de naissance", "dob", "birthdate", "birth_date"],
    "classe":         ["classe", "class", "level", "niveau"],
    "statut":         ["statut", "status", "etat", "état"],
}


def _normalize_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_").replace("-", "_")


def _map_headers(raw_headers: list[str]) -> dict[str, str]:
    """
    Retourne un dict {colonne_interne → header_original} pour les colonnes reconnues.
    """
    mapping: dict[str, str] = {}
    for raw in raw_headers:
        normalized = _normalize_header(raw)
        for internal, aliases in COL_ALIASES.items():
            if normalized in aliases and internal not in mapping:
                mapping[internal] = raw
                break
    return mapping


def _row_to_eleve_kwargs(row: dict[str, str], col_map: dict[str, str]) -> dict | None:
    """Convertit une ligne brute en kwargs pour créer un Eleve. Retourne None si ligne invalide."""
    def get(key: str) -> str:
        return (row.get(col_map.get(key, ""), "") or "").strip()

    nom    = get("nom")
    classe = get("classe")
    if not nom or not classe:
        return None
    return {
        "nom":            nom,
        "prenom":         get("prenom") or None,
        "genre":          get("genre") or None,
        "date_naissance": get("date_naissance") or None,
        "classe":         classe,
        "statut":         get("statut") or "actif",
    }


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_csv(content_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Extrait les lignes d'un fichier CSV. Retourne (rows, errors)."""
    text = content_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    raw_headers = list(reader.fieldnames or [])
    col_map = _map_headers(raw_headers)

    if "nom" not in col_map or "classe" not in col_map:
        raise HTTPException(
            status_code=422,
            detail=f"Colonnes requises : 'nom' et 'classe'. Colonnes détectées : {raw_headers}",
        )

    rows, errors = [], []
    for i, row in enumerate(reader, start=2):
        kwargs = _row_to_eleve_kwargs(dict(row), col_map)
        if kwargs is None:
            errors.append(f"Ligne {i} : 'nom' ou 'classe' manquant — ligne ignorée.")
        else:
            rows.append(kwargs)
    return rows, errors


def _parse_xlsx(content_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Extrait les lignes d'un fichier Excel .xlsx."""
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=422, detail="Le fichier Excel est vide.")

    # La première ligne non vide sert d'en-têtes
    raw_headers = [str(c) if c is not None else "" for c in all_rows[0]]
    col_map = _map_headers(raw_headers)

    if "nom" not in col_map or "classe" not in col_map:
        raise HTTPException(
            status_code=422,
            detail=f"Colonnes requises : 'nom' et 'classe'. Colonnes trouvées : {raw_headers}",
        )

    # Index → nom de colonne original
    idx_to_col = {raw_headers.index(v): v for v in col_map.values() if v in raw_headers}

    rows, errors = [], []
    for i, raw_row in enumerate(all_rows[1:], start=2):
        row_dict = {raw_headers[j]: str(v) if v is not None else "" for j, v in enumerate(raw_row) if j < len(raw_headers)}
        kwargs = _row_to_eleve_kwargs(row_dict, col_map)
        if kwargs is None:
            # Ligne complètement vide → on l'ignore silencieusement
            if all(not str(v or "").strip() for v in raw_row):
                continue
            errors.append(f"Ligne {i} : 'nom' ou 'classe' manquant — ligne ignorée.")
        else:
            rows.append(kwargs)
    wb.close()
    return rows, errors


def _parse_pdf(content_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Tente d'extraire un tableau d'élèves depuis un PDF avec pdfplumber.
    Cherche sur chaque page la première table contenant les colonnes requises.
    """
    rows, errors = [], []
    found = False

    with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                # La première ligne du tableau = en-têtes
                raw_headers = [str(c or "").strip() for c in table[0]]
                col_map = _map_headers(raw_headers)
                if "nom" not in col_map or "classe" not in col_map:
                    continue  # Ce tableau ne correspond pas
                found = True
                for i, raw_row in enumerate(table[1:], start=2):
                    row_dict = {raw_headers[j]: str(v or "").strip() for j, v in enumerate(raw_row) if j < len(raw_headers)}
                    kwargs = _row_to_eleve_kwargs(row_dict, col_map)
                    if kwargs is None:
                        if all(not v for v in row_dict.values()):
                            continue
                        errors.append(f"Page {page_num}, ligne {i} : 'nom' ou 'classe' manquant.")
                    else:
                        rows.append(kwargs)

    if not found:
        raise HTTPException(
            status_code=422,
            detail=(
                "Aucun tableau avec les colonnes 'nom' et 'classe' n'a été trouvé dans le PDF. "
                "Vérifiez que votre PDF contient bien un tableau structuré."
            ),
        )
    return rows, errors


# ── Bulk insert ───────────────────────────────────────────────────────────────

async def _bulk_insert(db, rows: list[dict]) -> int:
    """
    Insère les élèves en une seule requête executemany via pg_insert.
    Beaucoup plus rapide qu'une boucle de db.add() pour les gros volumes.
    ON CONFLICT DO NOTHING : les doublons exacts sont ignorés silencieusement.
    Retourne le nombre de lignes réellement insérées.
    """
    if not rows:
        return 0

    now = datetime.now(timezone.utc)
    bulk = [
        {
            "id":             uuid_module.uuid4(),
            "created_at":     now,
            "updated_at":     now,
            **kwargs,
        }
        for kwargs in rows
    ]

    stmt = pg_insert(Eleve).values(bulk).on_conflict_do_nothing()
    result = await db.execute(stmt)
    # rowcount peut être -1 sur certains drivers — on retourne len(bulk) par défaut
    return result.rowcount if result.rowcount >= 0 else len(bulk)


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("", response_model=Page[EleveResponse])
async def list_eleves(
    db: DB,
    _: AdminUser,
    page: Pagination,
    school_id:  Optional[uuid_module.UUID] = None,
    session_id: Optional[uuid_module.UUID] = None,
    classe:     Optional[str] = None,
    search:     Optional[str] = None,   # filtre nom/prénom
    ief:        Optional[str] = None,   # filtre par région (school.region)
) -> Page[EleveResponse]:
    from sqlalchemy.orm import joinedload
    from app.models.school import School as SchoolModel

    # Jointure école nécessaire si on filtre par IEF
    if ief:
        base = select(Eleve).join(SchoolModel, Eleve.school_id == SchoolModel.id)\
                            .where(SchoolModel.region == ief)\
                            .order_by(Eleve.nom)
    else:
        base = select(Eleve).order_by(Eleve.nom)

    if school_id:
        base = base.where(Eleve.school_id == school_id)
    if session_id:
        base = base.where(Eleve.session_id == session_id)
    if classe:
        base = base.where(Eleve.classe == classe)
    if search:
        term = f"%{search.strip()}%"
        base = base.where(or_(Eleve.nom.ilike(term), Eleve.prenom.ilike(term)))

    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    items_orm = (await db.execute(
        base.options(selectinload(Eleve.school)).offset(page.skip).limit(page.limit)
    )).scalars().all()

    items = [
        EleveResponse(
            id=e.id,
            nom=e.nom,
            prenom=e.prenom,
            classe=e.classe,
            genre=e.genre,
            date_naissance=e.date_naissance,
            statut=e.statut,
            school_id=e.school_id,
            session_id=e.session_id,
            school_name=e.school.name if e.school else None,
            school_region=e.school.region if e.school else None,
        )
        for e in items_orm
    ]
    return Page(total=total, skip=page.skip, limit=page.limit, items=items)


@router.post("", response_model=EleveResponse, status_code=status.HTTP_201_CREATED)
async def create_eleve(body: EleveCreate, db: DB, _: AdminUser) -> EleveResponse:
    eleve = Eleve(**body.model_dump())
    db.add(eleve)
    await db.flush()
    await db.refresh(eleve)
    return eleve


@router.patch("/{eleve_id}", response_model=EleveResponse)
async def update_eleve(eleve_id: uuid_module.UUID, body: EleveUpdate, db: DB, _: AdminUser) -> EleveResponse:
    result = await db.execute(select(Eleve).where(Eleve.id == eleve_id))
    eleve = result.scalar_one_or_none()
    if not eleve:
        raise HTTPException(status_code=404, detail="Élève introuvable.")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(eleve, field, value)
    await db.flush()
    await db.refresh(eleve)
    return eleve


@router.delete("/{eleve_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_eleve(eleve_id: uuid_module.UUID, db: DB, _: AdminUser) -> Response:
    result = await db.execute(select(Eleve).where(Eleve.id == eleve_id))
    eleve = result.scalar_one_or_none()
    if not eleve:
        raise HTTPException(status_code=404, detail="Élève introuvable.")
    await db.delete(eleve)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.delete("", status_code=status.HTTP_200_OK)
async def bulk_delete_eleves(
    db: DB,
    _: AdminUser,
    ids: list[uuid_module.UUID] = Body(..., embed=True),
) -> dict:
    """
    Supprime plusieurs élèves en une seule requête.
    Corps attendu : { "ids": ["uuid1", "uuid2", ...] }
    """
    if not ids:
        raise HTTPException(status_code=422, detail="La liste d'IDs ne peut pas être vide.")
    if len(ids) > 500:
        raise HTTPException(status_code=422, detail="Maximum 500 élèves supprimables en une seule opération.")

    result = await db.execute(
        delete(Eleve).where(Eleve.id.in_(ids))
    )
    return {"deleted": result.rowcount}


# ── Modèle de fichier (template) ──────────────────────────────────────────────

@router.get("/template/csv")
async def download_template_csv(_: AdminUser) -> StreamingResponse:
    """Retourne un fichier CSV vide avec les bons en-têtes pour guider l'import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nom", "prenom", "sexe", "date_naissance", "classe", "statut"])
    writer.writerow(["Diallo", "Aminata", "Fille", "2015-04-12", "CE2", "actif"])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=modele_eleves.csv"},
    )


@router.get("/template/xlsx")
async def download_template_xlsx(_: AdminUser) -> StreamingResponse:
    """Retourne un fichier Excel vide avec les bons en-têtes + une ligne exemple."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves"

    headers = ["nom", "prenom", "sexe", "date_naissance", "classe", "statut"]
    ws.append(headers)
    ws.append(["Diallo", "Aminata", "Fille", "2015-04-12", "CE2", "actif"])
    ws.append(["Ndiaye", "Ibrahima", "Garçon", "2014-09-03", "CM1", "actif"])

    # Largeurs de colonnes
    for i, col in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_eleves.xlsx"},
    )


# ── Export CSV ────────────────────────────────────────────────────────────────

@router.get("/export/csv")
async def export_eleves_csv(db: DB, _: AdminUser) -> StreamingResponse:
    items = (await db.execute(select(Eleve).order_by(Eleve.nom))).scalars().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nom", "prenom", "sexe", "date_naissance", "classe", "statut"])
    for e in items:
        writer.writerow([
            e.nom, e.prenom or "", e.genre or "", e.date_naissance or "",
            e.classe, e.statut,
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=eleves.csv"},
    )


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
async def export_eleves_xlsx(db: DB, _: AdminUser, fields: Optional[str] = None) -> StreamingResponse:
    items = (await db.execute(select(Eleve).order_by(Eleve.nom))).scalars().all()

    columns = [
        ("nom",     "Nom",               20),
        ("prenom",  "Prénom",            20),
        ("sexe",    "Sexe",              10),
        ("naiss",   "Date de naissance", 18),
        ("classe",  "Classe",            10),
        ("statut",  "Statut",            10),
    ]
    rows = [
        [
            e.nom,
            e.prenom or "",
            e.genre or "",
            e.date_naissance or "",
            e.classe,
            e.statut,
        ]
        for e in items
    ]

    return build_xlsx_response(
        sheet_title="Élèves",
        columns=columns,
        rows=rows,
        fields=fields,
        filename="eleves.xlsx",
    )


# ── Import unifié (CSV / Excel / PDF) ────────────────────────────────────────

@router.post("/import")
async def import_eleves(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Import d'élèves depuis un fichier CSV, Excel (.xlsx/.xls) ou PDF.
    Détection automatique du format selon l'extension et le content-type.
    Colonnes acceptées (insensible à la casse) :
      nom / prenom / sexe / date_naissance / classe / statut
    """
    content = await file.read()
    filename = (file.filename or "").lower()
    mime = (file.content_type or "").lower()

    # Détection du format
    if filename.endswith(".xlsx") or filename.endswith(".xls") or "spreadsheet" in mime or "excel" in mime:
        rows, errors = _parse_xlsx(content)
        fmt = "Excel"
    elif filename.endswith(".pdf") or mime == "application/pdf":
        rows, errors = _parse_pdf(content)
        fmt = "PDF"
    else:
        # Par défaut CSV (texte brut, .csv, .txt)
        rows, errors = _parse_csv(content)
        fmt = "CSV"

    imported = await _bulk_insert(db, rows)

    return {
        "format":   fmt,
        "imported": imported,
        "skipped":  len(errors),
        "errors":   errors[:50],  # max 50 erreurs retournées
    }


# ── Réimport depuis le fichier exporté ────────────────────────────────────────

@router.post("/reimport")
async def reimport_eleves_xlsx(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Réimporte le fichier Excel exporté depuis la plateforme.
    Format attendu : Nom, Prénom, Sexe, Date de naissance, Classe, Statut
    - Élève existant (même nom + prénom + classe) → ignoré
    - Nouvel élève → créé
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Impossible de lire le fichier Excel.")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 2:
        raise HTTPException(status_code=422, detail="Fichier vide ou sans données.")

    # ── Mapper les colonnes (comparaison lowercase directe) ───────────────────
    raw_header = [str(c) if c is not None else "" for c in all_rows[0]]
    header_lower = [h.strip().lower() for h in raw_header]

    HEADER_ALIASES: dict[str, list[str]] = {
        "nom":            ["nom", "name", "last_name", "lastname", "surname"],
        "prenom":         ["prenom", "prénom", "firstname", "first_name", "given_name"],
        "genre":          ["sexe", "genre", "sex", "gender"],
        "date_naissance": ["date naissance", "date_naissance", "naissance", "date de naissance", "dob", "birthdate"],
        "classe":         ["classe", "class", "level", "niveau"],
        "statut":         ["statut", "status", "etat", "état"],
    }
    col_idx: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(header_lower):
            if h in aliases and field not in col_idx:
                col_idx[field] = i
                break

    if "nom" not in col_idx or "classe" not in col_idx:
        raise HTTPException(
            status_code=422,
            detail=f"Colonnes 'nom' et 'classe' requises. Colonnes détectées : {raw_header}. Exportez d'abord depuis la plateforme.",
        )

    def col(row: tuple, name: str) -> str:
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ── Charger les élèves existants (1 requête) ──────────────────────────────
    existing_db = (await db.execute(select(Eleve))).scalars().all()
    existing_set: set[tuple[str, str, str]] = {
        (e.nom.strip().upper(), (e.prenom or "").strip().upper(), e.classe.strip().upper())
        for e in existing_db
    }

    rows_to_insert: list[dict] = []
    errors: list[str] = []
    skipped = 0

    for line_num, row in enumerate(all_rows[1:], start=2):
        if all(not str(v or "").strip() for v in row):
            continue

        nom    = col(row, "nom")
        classe = col(row, "classe")
        if not nom or not classe:
            errors.append(f"Ligne {line_num} : 'nom' ou 'classe' manquant — ignoré.")
            skipped += 1
            continue

        prenom = col(row, "prenom") or ""
        key = (nom.strip().upper(), prenom.strip().upper(), classe.strip().upper())
        if key in existing_set:
            skipped += 1
            continue

        rows_to_insert.append({
            "nom":            nom.strip(),
            "prenom":         prenom or None,
            "genre":          col(row, "genre") or None,
            "date_naissance": col(row, "date_naissance") or None,
            "classe":         classe.strip(),
            "statut":         col(row, "statut") or "actif",
        })
        existing_set.add(key)

    imported = await _bulk_insert(db, rows_to_insert)

    return {
        "imported": imported,
        "skipped":  skipped,
        "errors":   errors[:50],
    }


# ── Import CSV (rétrocompatibilité) ───────────────────────────────────────────

@router.post("/import/csv")
async def import_eleves_csv(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """Endpoint historique conservé pour compatibilité. Préférer /import."""
    content = await file.read()
    rows, errors = _parse_csv(content)
    imported = await _bulk_insert(db, rows)
    return {"imported": imported, "errors": errors}


# ── Import XLSX — format liste-élèves avec liaison école ─────────────────────

@router.post("/import/xlsx")
async def import_eleves_xlsx(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Importe les élèves depuis un fichier Excel au format liste-élèves ARED.
    Colonnes utilisées : name · Sexe · Classe · NIVEAU · SCHOOL
    Lie automatiquement chaque élève à son école via la colonne SCHOOL.
    Les écoles doivent exister au préalable (sinon l'élève est ignoré).
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Impossible de lire le fichier Excel.")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(status_code=422, detail="Fichier Excel vide.")

    # ── Mapper les colonnes ────────────────────────────────────────────────────
    raw_header = all_rows[0]
    header = [str(c).strip().lower() if c is not None else "" for c in raw_header]

    COL: dict[str, list[str]] = {
        "nom":    ["name", "nom", "last_name", "prenom_nom"],
        "prenom": ["prenom", "prénom", "firstname", "first_name"],
        "genre":  ["sexe", "genre", "sex", "gender"],
        "classe": ["classe", "class"],
        "niveau": ["niveau", "level"],
        "school": ["school", "école", "ecole"],
    }
    col_idx: dict[str, int] = {}
    for cname, aliases in COL.items():
        for alias in aliases:
            if alias in header:
                col_idx[cname] = header.index(alias)
                break

    if "nom" not in col_idx or "classe" not in col_idx:
        raise HTTPException(
            status_code=422,
            detail="Colonnes requises : 'name' (ou 'nom') et 'Classe'."
        )

    def col(row: tuple, name: str) -> str:
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ── Charger toutes les écoles existantes (1 requête) ──────────────────────
    from app.models.school import School as SchoolModel
    schools_db = (await db.execute(select(SchoolModel))).scalars().all()
    school_by_name: dict[str, uuid_module.UUID] = {
        s.name.upper(): s.id for s in schools_db
    }

    # ── Parser toutes les lignes ───────────────────────────────────────────────
    from datetime import datetime, timezone as _tz
    rows_to_insert: list[dict] = []
    errors: list[str] = []
    no_school_warned: set[str] = set()

    for i, row in enumerate(all_rows[1:], start=2):
        # Ligne vide → ignorer silencieusement
        if all(not str(v or "").strip() for v in row):
            continue

        nom    = col(row, "nom")
        classe = col(row, "classe")
        if not nom or not classe:
            errors.append(f"Ligne {i} : nom ou classe manquant — ignoré.")
            continue

        school_name  = col(row, "school")
        school_id    = school_by_name.get(school_name.upper()) if school_name else None
        if school_name and not school_id and school_name.upper() not in no_school_warned:
            no_school_warned.add(school_name.upper())
            errors.append(f"École '{school_name}' introuvable — élèves rattachés sans école.")

        rows_to_insert.append({
            "nom":       nom,
            "prenom":    col(row, "prenom") or None,
            "genre":     col(row, "genre")  or None,
            "classe":    classe,
            "statut":    "actif",
            "school_id": school_id,
        })

    imported = await _bulk_insert(db, rows_to_insert)

    return {
        "format":   "Excel (liste-élèves)",
        "imported": imported,
        "skipped":  len(errors),
        "errors":   errors[:50],
    }
