"""Endpoints Admin — Import / Export global des données du programme.

Export : un classeur Excel, une ligne par élève, reprenant toute la hiérarchie
(IEF → école → superviseur → tuteur → classe → élève), sur le même principe
que le fichier de référence BaseNWVFinale2026.

Import : réimporte ce même format. Comportement volontairement NON destructif —
chaque ligne est comparée à l'existant par une clé stable, et seuls les
enregistrements absents sont créés. Rien n'est jamais supprimé. Réimporter un
export intact ne crée donc rien : l'opération est idempotente.

Clés de correspondance
    École                : code_ecole, repli sur le nom
    Élève                : code_eleve, repli sur (école, classe, nom, prénom)
    Tuteur / superviseur : (école, nom) — insensible à la casse, aux accents et
                           aux espaces multiples. La paire inclut l'école parce
                           que deux personnes différentes peuvent porter le même
                           nom dans deux écoles différentes.

La seule mise à jour autorisée sur une fiche existante
    Ajouter une classe manquante à users.classes d'un tuteur. Sans elle, un
    élève ajouté dans une classe que le tuteur n'a pas encore serait bien créé
    en base mais resterait INVISIBLE dans l'app mobile : le rattachement
    élève → tuteur se fait sur (school_id, classe), et la classe manquerait du
    tableau (cf. CLAUDE.md §2). Les données ajoutées ne serviraient à rien.
"""
from __future__ import annotations

import io
import secrets
import unicodedata
import uuid as uuid_module
from typing import Optional

import openpyxl
from fastapi import APIRouter, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.core.deps import AdminUser, DB
from app.core.security import hash_password
from app.models.eleve import Eleve
from app.models.school import School
from app.models.school_classe import SchoolClasse
from app.models.user import User, UserRole
from app.schemas.user import _normalize_sn_phone

router = APIRouter(prefix="/import-export", tags=["Admin — Import / Export"])

COLUMNS = [
    "IEF", "Commune", "Code École", "École", "Langue", "Directeur", "Téléphone Directeur",
    "Superviseur", "Téléphone Superviseur",
    "Tuteur", "Téléphone Tuteur",
    "Classe", "Niveau",
    "Code Élève", "Nom", "Prénom", "Sexe",
    "Type", "Groupe Lecture", "Groupe Maths", "Statut",
]


# ── Export ───────────────────────────────────────────────────────────────────

@router.get("/export")
async def export_all(db: DB, _: AdminUser) -> StreamingResponse:
    schools = (await db.execute(select(School))).scalars().all()
    users = (await db.execute(
        select(User).where(User.role.in_((UserRole.enseignant, UserRole.superviseur)))
    )).scalars().all()
    classes = (await db.execute(select(SchoolClasse))).scalars().all()
    eleves = (await db.execute(select(Eleve).options(selectinload(Eleve.school)))).scalars().all()

    teachers_by_school: dict[uuid_module.UUID, list[User]] = {}
    supervisors_by_school: dict[uuid_module.UUID, list[User]] = {}
    for u in users:
        if u.school_id is None:
            continue
        bucket = teachers_by_school if u.role == UserRole.enseignant else supervisors_by_school
        bucket.setdefault(u.school_id, []).append(u)

    niveau_by_class: dict[tuple[uuid_module.UUID, str], str] = {
        (c.school_id, c.name): c.niveau for c in classes
    }

    def tuteur_for(school_id: Optional[uuid_module.UUID], classe: str) -> Optional[User]:
        if school_id is None:
            return None
        for t in teachers_by_school.get(school_id, []):
            if classe in (t.classes or []):
                return t
        teachers = teachers_by_school.get(school_id, [])
        return teachers[0] if len(teachers) == 1 else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for e in sorted(eleves, key=lambda x: (x.school.name if x.school else "", x.classe, x.nom)):
        school = e.school
        sup_list = supervisors_by_school.get(school.id, []) if school else []
        sup_name = " | ".join(s.name for s in sup_list) if sup_list else None
        tut = tuteur_for(school.id, e.classe) if school else None
        niveau = niveau_by_class.get((school.id, e.classe)) if school else None

        ws.append([
            school.region if school else None,
            school.city if school else None,
            school.code_ecole if school else None,
            school.name if school else None,
            school.langue if school else None,
            school.director if school else None,
            school.director_phone if school else None,
            sup_name,
            (sup_list[0].phone if sup_list else None),
            tut.name if tut else None,
            tut.phone if tut else None,
            e.classe,
            niveau,
            e.code_eleve,
            e.nom,
            e.prenom,
            e.genre,
            e.statut_selection,
            e.groupe_lecture,
            e.groupe_maths,
            e.statut,
        ])

    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(12, min(28, len(col) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=export_ndawwune.xlsx"},
    )


# ── Import ───────────────────────────────────────────────────────────────────

class ImportSummary(BaseModel):
    ecoles_creees:      int = 0
    superviseurs_crees: int = 0
    tuteurs_crees:      int = 0
    classes_creees:     int = 0
    classes_ajoutees_tuteurs: int = 0  # classes ajoutées à un tuteur existant
    eleves_crees:       int = 0
    ignores:            int = 0  # lignes dont l'élève existait déjà
    erreurs:            list[str] = []


def _s(v) -> Optional[str]:
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def _nkey(v) -> str:
    """Clé de correspondance d'un nom de personne : insensible à la casse, aux
    accents et aux espaces multiples.

    Les noms saisis à la main dans un fichier réimporté ne portent pas toujours
    les accents du dashboard (« SATOU SENE » pour « SATOU SÉNE »), et certaines
    fiches ont un double espace (« MODOU  FALL »). Sans cette neutralisation, un
    nom déjà connu est vu comme nouveau et un compte en double est créé.
    """
    if v is None:
        return ""
    s = unicodedata.normalize("NFD", str(v))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return " ".join(s.split()).upper()


def _unique_placeholder_phone(taken: set[str]) -> str:
    """Numéro à 9 chiffres garanti libre, pour un compte créé sans contact fourni
    dans le fichier — doit être changé manuellement ensuite."""
    while True:
        candidate = "90" + "".join(secrets.choice("0123456789") for _ in range(7))
        if candidate not in taken:
            taken.add(candidate)
            return candidate


@router.post("/import", response_model=ImportSummary)
async def import_all(db: DB, _: AdminUser, file: UploadFile = File(...)) -> ImportSummary:
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return ImportSummary(erreurs=["Impossible de lire le fichier Excel."])
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return ImportSummary(erreurs=["Fichier vide."])

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: header.index(name) for name in COLUMNS if name in header}
    missing = [c for c in ("École", "Classe", "Nom") if c not in idx]
    if missing:
        return ImportSummary(erreurs=[f"Colonnes requises manquantes : {missing}. "
                                       f"Réexportez depuis « Import / Export » pour garantir le bon format."])

    def get(row: tuple, name: str):
        i = idx.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    summary = ImportSummary()

    # ── Écoles existantes (clé : code_ecole, repli sur nom normalisé) ─────────
    schools_db = (await db.execute(select(School))).scalars().all()
    schools_by_code = {s.code_ecole: s for s in schools_db if s.code_ecole is not None}
    schools_by_name = {s.name.strip().upper(): s for s in schools_db}

    users_db = (await db.execute(select(User))).scalars().all()
    # Deux personnes différentes peuvent porter le même nom dans deux écoles
    # différentes (c'est le cas en base pour AISSATA NDIAYE). La clé de
    # correspondance est donc (école, nom, rôle) ; on retombe sur les comptes
    # sans école, puis on crée.
    users_by_school: dict[tuple[Optional[uuid_module.UUID], str, UserRole], User] = {}
    users_sans_ecole: dict[tuple[str, UserRole], User] = {}
    for u in users_db:
        users_by_school[(u.school_id, _nkey(u.name), u.role)] = u
        if u.school_id is None:
            users_sans_ecole[(_nkey(u.name), u.role)] = u
    taken_phones = {u.phone for u in users_db if u.phone}

    def trouver_compte(school_id, nom: str, role: UserRole) -> Optional[User]:
        u = users_by_school.get((school_id, _nkey(nom), role))
        if u is not None:
            return u
        u = users_sans_ecole.get((_nkey(nom), role))
        if u is not None:            # compte orphelin : on le rattache à l'école
            u.school_id = school_id
            users_by_school[(school_id, _nkey(nom), role)] = u
            users_sans_ecole.pop((_nkey(nom), role), None)
        return u

    def enregistrer_compte(u: User) -> None:
        users_by_school[(u.school_id, _nkey(u.name), u.role)] = u

    def ajouter_classe(tut: User, classe: str, niveau: Optional[str]) -> bool:
        """Complète users.classes / users.niveau d'un enseignant.

        Sans cela, un élève ajouté dans une classe que le tuteur n'a pas encore
        est créé en base mais reste INVISIBLE dans l'app : le rattachement se
        fait sur (school_id, classe) et la classe manquerait du tableau. C'est
        la seule mise à jour que l'import s'autorise sur une fiche existante,
        parce que sans elle les données ajoutées ne servent à rien.
        """
        change = False
        if classe not in (tut.classes or []):
            tut.classes = sorted([*(tut.classes or []), classe])
            change = True
        if niveau and niveau not in (tut.niveau or []):
            tut.niveau = sorted([*(tut.niveau or []), niveau])
            change = True
        return change

    classes_db = (await db.execute(select(SchoolClasse))).scalars().all()
    classes_by_key = {(c.school_id, c.name): c for c in classes_db}

    eleves_db = (await db.execute(select(Eleve))).scalars().all()
    eleves_by_code = {e.code_eleve: e for e in eleves_db if e.code_eleve}
    eleves_by_identity = {
        (e.school_id, e.classe, (e.nom or "").strip().upper(), (e.prenom or "").strip().upper()): e
        for e in eleves_db
    }

    for line_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            ecole_name = _s(get(row, "École"))
            classe = _s(get(row, "Classe"))
            nom_eleve = _s(get(row, "Nom"))
            if not ecole_name or not classe or not nom_eleve:
                summary.erreurs.append(f"Ligne {line_num} : École, Classe et Nom sont requis — ignorée.")
                continue

            # 1) École
            code_ecole_raw = get(row, "Code École")
            code_ecole = int(code_ecole_raw) if code_ecole_raw not in (None, "") else None
            school = (schools_by_code.get(code_ecole) if code_ecole is not None else None) \
                or schools_by_name.get(ecole_name.upper())
            if school is None:
                school = School(
                    id=uuid_module.uuid4(),
                    name=ecole_name,
                    code_ecole=code_ecole,
                    region=_s(get(row, "IEF")),
                    city=_s(get(row, "Commune")),
                    langue=(_s(get(row, "Langue")) or "").lower() or None,
                    director=_s(get(row, "Directeur")),
                    director_phone=_normalize_sn_phone(_s(get(row, "Téléphone Directeur"))),
                )
                db.add(school)
                await db.flush()
                if code_ecole is not None:
                    schools_by_code[code_ecole] = school
                schools_by_name[ecole_name.upper()] = school
                summary.ecoles_creees += 1

            # 2) Superviseur (créé seulement s'il n'existe pas déjà)
            sup_name = _s(get(row, "Superviseur"))
            if sup_name:
                for one_name in [n.strip() for n in sup_name.split("|") if n.strip()]:
                    if trouver_compte(school.id, one_name, UserRole.superviseur) is None:
                        phone = _normalize_sn_phone(_s(get(row, "Téléphone Superviseur"))) \
                            or _unique_placeholder_phone(taken_phones)
                        new_sup = User(
                            id=uuid_module.uuid4(), name=one_name,
                            phone=phone,
                            password_hash=hash_password(secrets.token_urlsafe(12)),
                            role=UserRole.superviseur, school_id=school.id,
                            must_change_password=True,
                        )
                        db.add(new_sup)
                        await db.flush()
                        enregistrer_compte(new_sup)
                        taken_phones.add(phone)
                        summary.superviseurs_crees += 1

            # 3) Tuteur / enseignant
            niveau = _s(get(row, "Niveau")) or classe.split()[0]
            tut_name = _s(get(row, "Tuteur"))
            if tut_name:
                tut = trouver_compte(school.id, tut_name, UserRole.enseignant)
                if tut is None:
                    phone = _normalize_sn_phone(_s(get(row, "Téléphone Tuteur"))) \
                        or _unique_placeholder_phone(taken_phones)
                    tut = User(
                        id=uuid_module.uuid4(), name=tut_name,
                        phone=phone,
                        password_hash=hash_password(secrets.token_urlsafe(12)),
                        role=UserRole.enseignant, school_id=school.id,
                        classes=[classe], niveau=[niveau],
                        must_change_password=True,
                    )
                    db.add(tut)
                    await db.flush()
                    enregistrer_compte(tut)
                    taken_phones.add(phone)
                    summary.tuteurs_crees += 1
                elif ajouter_classe(tut, classe, niveau):
                    # Fiche conservée telle quelle, seule la liste des classes
                    # s'enrichit — sinon les élèves de cette classe seraient
                    # créés mais invisibles pour leur tuteur.
                    await db.flush()
                    summary.classes_ajoutees_tuteurs += 1

            # 4) Classe (school_classes)
            class_key = (school.id, classe)
            if class_key not in classes_by_key:
                new_classe = SchoolClasse(id=uuid_module.uuid4(), school_id=school.id, name=classe, niveau=niveau)
                db.add(new_classe)
                await db.flush()
                classes_by_key[class_key] = new_classe
                summary.classes_creees += 1

            # 5) Élève (clé : code_eleve, repli sur identité)
            code_eleve = _s(get(row, "Code Élève"))
            prenom = _s(get(row, "Prénom"))
            identity_key = (school.id, classe, nom_eleve.upper(), (prenom or "").upper())
            existing_eleve = (eleves_by_code.get(code_eleve) if code_eleve else None) \
                or eleves_by_identity.get(identity_key)
            if existing_eleve is None:
                new_eleve = Eleve(
                    id=uuid_module.uuid4(),
                    nom=nom_eleve, prenom=prenom, code_eleve=code_eleve,
                    classe=classe, genre=_s(get(row, "Sexe")),
                    statut=_s(get(row, "Statut")) or "actif",
                    statut_selection=_s(get(row, "Type")),
                    groupe_lecture=_s(get(row, "Groupe Lecture")),
                    groupe_maths=_s(get(row, "Groupe Maths")),
                    school_id=school.id,
                )
                db.add(new_eleve)
                await db.flush()
                if code_eleve:
                    eleves_by_code[code_eleve] = new_eleve
                eleves_by_identity[identity_key] = new_eleve
                summary.eleves_crees += 1
            else:
                # Une ligne = un élève : c'est la seule unité de comptage qui
                # ait un sens pour l'utilisateur. Compter aussi l'école, le
                # tuteur et la classe donnait ~5 « ignorés » par ligne.
                summary.ignores += 1

        except Exception as exc:  # une ligne en erreur ne doit pas bloquer les autres
            summary.erreurs.append(f"Ligne {line_num} : {exc}")

    summary.erreurs = summary.erreurs[:50]
    return summary
