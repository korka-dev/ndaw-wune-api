import csv
import io
import uuid
from typing import Optional
import openpyxl
from fastapi import APIRouter, File, HTTPException, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select

from app.core.deps import AdminUser, DB
from app.core.export_utils import build_xlsx_response
from app.core.pagination import Page, Pagination
from app.models.school import School
from app.schemas.school import SchoolCreate, SchoolUpdate, SchoolResponse

router = APIRouter(prefix="/schools", tags=["Admin — Écoles"])


async def _check_phone_unique(db, phone: str | None, exclude_id: uuid.UUID | None = None) -> None:
    """Lève 409 si le numéro de téléphone est déjà utilisé par une autre école."""
    if not phone:
        return
    q = select(School).where(School.director_phone == phone)
    if exclude_id:
        q = q.where(School.id != exclude_id)
    existing = await db.scalar(q)
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Le numéro {phone} est déjà associé à l'école « {existing.name} ».",
        )


@router.get("", response_model=Page[SchoolResponse])
async def list_schools(db: DB, _: AdminUser, page: Pagination) -> Page[SchoolResponse]:
    base  = select(School).order_by(School.name)
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    items = (await db.execute(base.offset(page.skip).limit(page.limit))).scalars().all()
    return Page(total=total, skip=page.skip, limit=page.limit, items=items)


@router.post("", response_model=SchoolResponse, status_code=status.HTTP_201_CREATED)
async def create_school(body: SchoolCreate, db: DB, _: AdminUser):
    await _check_phone_unique(db, body.director_phone)
    school = School(**body.model_dump())
    db.add(school)
    await db.commit()
    await db.refresh(school)
    return school


# ── Export CSV — DOIT être avant /{school_id} pour éviter qu'FastAPI confonde "export" avec un UUID ──

@router.get("/export/csv")
async def export_schools_csv(db: DB, _: AdminUser) -> StreamingResponse:
    items = (await db.execute(select(School).order_by(School.name))).scalars().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nom", "region", "commune", "directeur", "telephone_directeur"])
    for s in items:
        writer.writerow([s.name, s.region or "", s.city or "", s.director or "", s.director_phone or ""])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ecoles.csv"},
    )


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
async def export_schools_xlsx(db: DB, _: AdminUser, fields: Optional[str] = None) -> StreamingResponse:
    items = (await db.execute(select(School).order_by(School.name))).scalars().all()

    columns = [
        ("nom",       "Nom de l'école",   30),
        ("region",    "Région (IEF)",     20),
        ("commune",   "Commune / Ville",  20),
        ("directeur", "Directeur(trice)", 25),
        ("telephone", "Téléphone",        18),
    ]
    rows = [
        [
            s.name,
            s.region or "",
            s.city or "",
            s.director or "",
            s.director_phone or "",
        ]
        for s in items
    ]

    return build_xlsx_response(
        sheet_title="Écoles",
        columns=columns,
        rows=rows,
        fields=fields,
        filename="ecoles.xlsx",
    )


# ── Import CSV ────────────────────────────────────────────────────────────────

@router.post("/import/csv")
async def import_schools_csv(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    if "nom" not in {f.lower().strip() for f in (reader.fieldnames or [])}:
        raise HTTPException(status_code=422, detail="Colonne 'nom' requise.")
    imported = 0
    errors: list[str] = []
    for i, row in enumerate(reader, start=2):
        nom = (row.get("nom") or "").strip()
        if not nom:
            errors.append(f"Ligne {i} : nom manquant.")
            continue
        school = School(
            name=nom,
            region=(row.get("region") or "").strip() or None,
            city=(row.get("commune") or row.get("city") or "").strip() or None,
            director=(row.get("directeur") or row.get("director") or "").strip() or None,
            director_phone=(row.get("telephone_directeur") or row.get("director_phone") or "").strip() or None,
        )
        db.add(school)
        imported += 1
    if imported:
        await db.commit()
    return {"imported": imported, "errors": errors}


# ── Réimport depuis le fichier exporté ────────────────────────────────────────

@router.post("/reimport")
async def reimport_schools_xlsx(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Réimporte le fichier Excel exporté depuis la plateforme.
    Format attendu : Nom de l'école, Région (IEF), Commune / Ville, Directeur(trice), Téléphone
    - École existante (même nom) → mise à jour des champs modifiés
    - Nouvelle école → créée
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Impossible de lire le fichier Excel.")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 2:
        raise HTTPException(status_code=422, detail="Fichier vide ou sans données.")

    # ── Mapper les colonnes (comparaison lowercase directe) ───────────────────
    raw_header = [str(c) if c is not None else "" for c in all_rows[0]]
    header_lower = [h.strip().lower() for h in raw_header]

    HEADER_ALIASES: dict[str, list[str]] = {
        "nom":       ["nom de l'école", "nom de l ecole", "nom", "name", "école", "ecole", "school"],
        "region":    ["région (ief)", "region (ief)", "région", "region", "ief"],
        "commune":   ["commune / ville", "commune/ville", "commune", "ville", "city"],
        "directeur": ["directeur(trice)", "directeur", "director"],
        "phone":     ["téléphone", "telephone", "phone", "tel"],
    }
    col_idx: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(header_lower):
            if h in aliases and field not in col_idx:
                col_idx[field] = i
                break

    if "nom" not in col_idx:
        raise HTTPException(
            status_code=422,
            detail=f"Colonne 'Nom de l\\'école' introuvable. Exportez d'abord depuis la plateforme. Colonnes détectées : {raw_header}",
        )

    def col(row: tuple, name: str) -> str:
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ── Charger les écoles et téléphones existants (1 requête) ────────────────
    all_schools_db = (await db.execute(select(School))).scalars().all()
    school_by_name: dict[str, School] = {s.name.strip().upper(): s for s in all_schools_db}
    phone_to_id: dict[str, str] = {
        s.director_phone: str(s.id)
        for s in all_schools_db
        if s.director_phone
    }

    updated = 0
    created = 0
    skipped = 0
    errors: list[str] = []

    for line_num, row in enumerate(all_rows[1:], start=2):
        if all(not str(v or "").strip() for v in row):
            continue

        nom = col(row, "nom")
        if not nom:
            skipped += 1
            continue

        region    = col(row, "region")    or None
        commune   = col(row, "commune")   or None
        directeur = col(row, "directeur") or None
        phone     = col(row, "phone")     or None

        key = nom.strip().upper()
        existing = school_by_name.get(key)

        # Vérification unicité téléphone
        if phone:
            conflict_id = phone_to_id.get(phone)
            if conflict_id and (existing is None or conflict_id != str(existing.id)):
                errors.append(f"Ligne {line_num} — {nom} : le téléphone {phone} est déjà utilisé par une autre école.")
                skipped += 1
                continue

        if existing:
            changed = False
            if region is not None and existing.region != region:
                existing.region = region; changed = True
            if commune is not None and existing.city != commune:
                existing.city = commune; changed = True
            if directeur is not None and existing.director != directeur:
                existing.director = directeur; changed = True
            if phone and existing.director_phone != phone:
                if existing.director_phone:
                    phone_to_id.pop(existing.director_phone, None)
                existing.director_phone = phone
                phone_to_id[phone] = str(existing.id)
                changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            school = School(
                name=nom.strip(),
                region=region,
                city=commune,
                director=directeur,
                director_phone=phone,
            )
            db.add(school)
            await db.flush()
            await db.refresh(school)
            school_by_name[key] = school
            if phone:
                phone_to_id[phone] = str(school.id)
            created += 1

    return {
        "updated": updated,
        "created": created,
        "skipped": skipped,
        "errors":  errors,
    }


# ── Routes paramétrées — après les routes fixes ──────────────────────────────

@router.get("/{school_id}", response_model=SchoolResponse)
async def get_school(school_id: uuid.UUID, db: DB, _: AdminUser):
    result = await db.execute(select(School).where(School.id == school_id))
    school = result.scalar_one_or_none()
    if not school:
        raise HTTPException(status_code=404, detail="École introuvable.")
    return school


@router.patch("/{school_id}", response_model=SchoolResponse)
async def update_school(school_id: uuid.UUID, body: SchoolUpdate, db: DB, _: AdminUser):
    result = await db.execute(select(School).where(School.id == school_id))
    school = result.scalar_one_or_none()
    if not school:
        raise HTTPException(status_code=404, detail="École introuvable.")
    await _check_phone_unique(db, body.director_phone, exclude_id=school_id)
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(school, field, value)
    await db.commit()
    await db.refresh(school)
    return school


@router.delete("/{school_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_school(school_id: uuid.UUID, db: DB, _: AdminUser) -> Response:
    result = await db.execute(select(School).where(School.id == school_id))
    school = result.scalar_one_or_none()
    if not school:
        raise HTTPException(status_code=404, detail="École introuvable.")
    await db.delete(school)
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ── Import XLSX — format liste-élèves (IEF · COMMUNE · SCHOOL · …) ───────────

@router.post("/import/xlsx")
async def import_schools_xlsx(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Importe les écoles depuis un fichier Excel au format liste-élèves.
    Colonnes utilisées : IEF · COMMUNE · SCHOOL
    Les autres colonnes (enseignant, élève, etc.) sont ignorées.
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Impossible de lire le fichier Excel.")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(status_code=422, detail="Fichier Excel vide.")

    # ── Mapper les colonnes ────────────────────────────────────────────────────
    raw_header = all_rows[0]
    header = [str(c).strip().lower() if c is not None else "" for c in raw_header]

    COL: dict[str, list[str]] = {
        "school":  ["school", "école", "ecole", "nom_ecole"],
        "ief":     ["ief"],
        "commune": ["commune", "city", "ville"],
    }
    col_idx: dict[str, int] = {}
    for name, aliases in COL.items():
        for alias in aliases:
            if alias in header:
                col_idx[name] = header.index(alias)
                break

    if "school" not in col_idx:
        raise HTTPException(
            status_code=422,
            detail="Colonne 'SCHOOL' (nom de l'école) introuvable dans le fichier."
        )

    def col(row: tuple, name: str) -> str:
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ── Agréger les écoles uniques ─────────────────────────────────────────────
    schools_map: dict[str, dict] = {}  # key = nom.upper()
    for row in all_rows[1:]:
        name_raw = col(row, "school")
        if not name_raw:
            continue
        key = name_raw.upper()
        if key not in schools_map:
            schools_map[key] = {
                "name":   name_raw,
                "region": col(row, "ief") or None,
                "city":   col(row, "commune") or None,
            }

    if not schools_map:
        raise HTTPException(status_code=422, detail="Aucune école trouvée dans le fichier.")

    # ── Charger les écoles existantes ──────────────────────────────────────────
    existing = (await db.execute(select(School))).scalars().all()
    existing_names: set[str] = {s.name.upper() for s in existing}

    # ── Créer les nouvelles écoles en lot ──────────────────────────────────────
    imported = 0
    skipped  = 0
    for key, info in schools_map.items():
        if key in existing_names:
            skipped += 1
            continue
        db.add(School(name=info["name"], region=info["region"], city=info["city"]))
        existing_names.add(key)  # éviter doublons dans le même batch
        imported += 1

    if imported:
        await db.commit()

    return {"imported": imported, "skipped": skipped}
