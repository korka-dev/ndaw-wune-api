from __future__ import annotations

import csv
import io
import uuid

import openpyxl
from fastapi import APIRouter, Body, File, HTTPException, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, or_, select

from app.core.deps import AdminUser, DB
from app.core.pagination import Page, Pagination
from app.core.security import hash_password
from app.models.school import School
from app.models.user import User, UserRole, UserStatus
from app.schemas.user import UserCreate, UserResponse, UserUpdate
from app.services import user_service

router = APIRouter(prefix="/teachers", tags=["Admin — Enseignants"])


@router.get("", response_model=Page[UserResponse])
async def list_teachers(db: DB, _: AdminUser, page: Pagination) -> Page[UserResponse]:
    total, items = await user_service.list_by_role(db, UserRole.enseignant, page.skip, page.limit)
    return Page(total=total, skip=page.skip, limit=page.limit, items=items)


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_teacher(body: UserCreate, db: DB, _: AdminUser) -> UserResponse:
    # force_role garantit que le rôle est toujours "enseignant", peu importe la valeur envoyée
    return await user_service.create_user(db, body, force_role=UserRole.enseignant)


# ── Export CSV — DOIT être avant /{teacher_id} pour éviter qu'FastAPI confonde "export" avec un UUID ──

@router.get("/export/csv")
async def export_teachers_csv(db: DB, _: AdminUser) -> StreamingResponse:
    items = (await db.execute(
        select(User).where(User.role == UserRole.enseignant).order_by(User.name)
    )).scalars().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nom", "telephone", "titre", "email", "niveau", "classes"])
    for t in items:
        writer.writerow([
            t.name, t.phone or "", t.title or "", t.email or "",
            "|".join(t.niveau or []),
            "|".join(t.classes or []),
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=enseignants.csv"},
    )


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
async def export_teachers_xlsx(db: DB, _: AdminUser) -> StreamingResponse:
    items = (await db.execute(
        select(User).where(User.role == UserRole.enseignant).order_by(User.name)
    )).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enseignants"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1e6fbf", end_color="1e6fbf", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Nom Complet", "Téléphone", "Titre / Fonction", "Email", "Niveaux", "Classes", "Statut"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for t in items:
        ws.append([
            t.name,
            t.phone or "",
            t.title or "",
            t.email or "",
            ", ".join(t.niveau or []),
            ", ".join(t.classes or []),
            t.status,
        ])

    col_widths = [25, 18, 20, 25, 20, 20, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enseignants.xlsx"},
    )


# ── Import CSV ────────────────────────────────────────────────────────────────

@router.post("/import/csv")
async def import_teachers_csv(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = {f.lower().strip() for f in (reader.fieldnames or [])}
    if "nom" not in headers:
        raise HTTPException(status_code=422, detail="Colonne 'nom' requise.")

    imported = 0
    errors: list[str] = []
    for i, row in enumerate(reader, start=2):
        nom = (row.get("nom") or "").strip()
        phone = (row.get("telephone") or row.get("phone") or "").strip() or None
        email = (row.get("email") or "").strip() or None
        if not nom or (not phone and not email):
            errors.append(f"Ligne {i} : nom + (téléphone ou email) requis.")
            continue
        niveau_raw  = (row.get("niveau") or "").strip()
        niveau  = [n.strip() for n in niveau_raw.split("|")  if n.strip()] if niveau_raw  else []
        classes_raw = (row.get("classes") or "").strip()
        classes = [c.strip() for c in classes_raw.split("|") if c.strip()] if classes_raw else []
        body = UserCreate(name=nom, phone=phone, email=email,
                          title=(row.get("titre") or "").strip() or None,
                          role=UserRole.enseignant,
                          niveau=niveau or None,
                          classes=classes or None)
        await user_service.create_user(db, body, force_role=UserRole.enseignant)
        imported += 1
    return {"imported": imported, "errors": errors}


# ── Import XLSX — format liste élèves (IEF, COMMUNE, SCHOOL, enseignant, NIVEAU, Classe, ...) ──

_DEFAULT_PW_HASH = hash_password("P@sser123")  # calculé une seule fois au démarrage

@router.post("/import/xlsx")
async def import_teachers_xlsx(db: DB, _: AdminUser, file: UploadFile = File(...)) -> dict:
    """
    Importe les enseignants depuis un fichier Excel au format liste-élèves.
    Colonnes attendues : IEF · COMMUNE · SCHOOL · enseignant · NIVEAU · Classe
    (les colonnes élève name / Sexe sont ignorées)

    Stratégie batch pour tenir sur N lignes :
      1. Lecture complète du fichier en mémoire (openpyxl read_only)
      2. Agrégation Python pure → dict des enseignants uniques
      3. Création des écoles manquantes en lot (1 flush)
      4. Chargement des enseignants existants en 1 requête
      5. Création des nouveaux enseignants en lot (flush toutes les 200 entités)
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Impossible de lire le fichier Excel.")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(status_code=422, detail="Fichier Excel vide.")

    # ── Mapper les colonnes ────────────────────────────────────────────────────
    raw_header = all_rows[0]
    header = [str(c).strip().lower() if c is not None else "" for c in raw_header]

    COL_ALIASES: dict[str, list[str]] = {
        "ief":        ["ief"],
        "commune":    ["commune"],
        "school":     ["school", "école", "ecole"],
        "enseignant": ["enseignant", "teacher", "prof"],
        "niveau":     ["niveau", "level", "niveaux"],
        "classe":     ["classe", "class"],
    }

    # Pré-calculer les indices une seule fois
    col_idx: dict[str, int] = {}
    for name, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias in header:
                col_idx[name] = header.index(alias)
                break

    def col(row: tuple, name: str) -> str:
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ── Étape 1 : Agrégation pure Python — O(n) sur toutes les lignes ─────────
    teachers_map: dict[tuple[str, str], dict] = {}
    for row in all_rows[1:]:
        t_name = col(row, "enseignant")
        s_name = col(row, "school")
        if not t_name or not s_name:
            continue
        key = (s_name.upper(), t_name.upper())
        if key not in teachers_map:
            teachers_map[key] = {
                "name":    t_name.title(),
                "school":  s_name,
                "ief":     col(row, "ief"),
                "commune": col(row, "commune"),
                "niveaux": set(),
                "classes": set(),
            }
        niv = col(row, "niveau")
        cls = col(row, "classe")
        if niv:
            teachers_map[key]["niveaux"].add(niv)
        if cls:
            teachers_map[key]["classes"].add(cls)

    if not teachers_map:
        raise HTTPException(
            status_code=422,
            detail=(
                "Aucun enseignant trouvé. Vérifiez que le fichier contient "
                "les colonnes 'enseignant' et 'SCHOOL'."
            ),
        )

    # ── Étape 2 : Charger toutes les écoles existantes (1 requête) ────────────
    schools_db = (await db.execute(select(School))).scalars().all()
    school_by_name: dict[str, School] = {s.name.upper(): s for s in schools_db}

    # ── Étape 3 : Créer les écoles manquantes en LOT ──────────────────────────
    new_schools: list[School] = []
    needed_school_names = {info["school"].upper() for info in teachers_map.values()}
    for name_upper in needed_school_names:
        if name_upper not in school_by_name:
            # Trouver l'info depuis le premier enseignant de cette école
            info = next(v for k, v in teachers_map.items() if k[0] == name_upper)
            school_obj = School(
                name=info["school"],
                region=info["ief"] or None,
                city=info["commune"] or None,
            )
            db.add(school_obj)
            new_schools.append(school_obj)

    if new_schools:
        await db.flush()  # 1 seul flush pour toutes les nouvelles écoles
        for s in new_schools:
            await db.refresh(s)
            school_by_name[s.name.upper()] = s

    schools_created = len(new_schools)

    # ── Étape 4 : Charger les enseignants existants en 1 requête ──────────────
    school_ids = [s.id for s in school_by_name.values()]
    existing_teachers = (await db.execute(
        select(User.name, User.school_id).where(
            User.role == UserRole.enseignant,
            User.school_id.in_(school_ids),
        )
    )).all()
    existing_set: set[tuple[str, str]] = {
        (row.name.upper(), str(row.school_id)) for row in existing_teachers
    }

    # ── Étape 5 : Créer les enseignants manquants en LOT (flush / 200) ────────
    imported  = 0
    skipped   = 0
    errors: list[str] = []
    batch_size = 200
    pending    = 0

    for (school_upper, teacher_upper), info in teachers_map.items():
        school_obj = school_by_name.get(school_upper)
        if school_obj is None:
            errors.append(f"{info['name']} : école '{info['school']}' introuvable après création.")
            continue

        if (teacher_upper, str(school_obj.id)) in existing_set:
            skipped += 1
            continue

        try:
            user = User(
                name=info["name"],
                role=UserRole.enseignant,
                status=UserStatus.actif,
                password_hash=_DEFAULT_PW_HASH,
                must_change_password=True,
                school_id=school_obj.id,
                niveau=sorted(info["niveaux"]) if info["niveaux"] else None,
                classes=sorted(info["classes"]) if info["classes"] else None,
            )
            db.add(user)
            # Marquer comme vu pour éviter les doublons dans le même batch
            existing_set.add((teacher_upper, str(school_obj.id)))
            pending += 1
            imported += 1

            if pending >= batch_size:
                await db.flush()
                pending = 0

        except Exception as exc:
            errors.append(f"{info['name']} ({info['school']}) : {exc}")

    if pending:
        await db.flush()  # flush du dernier batch

    return {
        "imported":        imported,
        "skipped":         skipped,
        "schools_created": schools_created,
        "errors":          errors,
    }


# ── Routes paramétrées — après les routes fixes ──────────────────────────────

@router.get("/{teacher_id}", response_model=UserResponse)
async def get_teacher(teacher_id: uuid.UUID, db: DB, _: AdminUser) -> UserResponse:
    return await user_service.get_by_id(db, teacher_id)


@router.patch("/{teacher_id}", response_model=UserResponse)
async def update_teacher(teacher_id: uuid.UUID, body: UserUpdate, db: DB, _: AdminUser) -> UserResponse:
    return await user_service.update_user(db, teacher_id, body)


@router.delete("", status_code=status.HTTP_200_OK)
async def bulk_delete_teachers(
    db: DB,
    current_user: AdminUser,
    ids: list[uuid.UUID] = Body(..., embed=True),
) -> dict:
    """
    Supprime plusieurs enseignants en une seule requête.
    Corps attendu : { "ids": ["uuid1", "uuid2", ...] }
    """
    if not ids:
        raise HTTPException(status_code=422, detail="La liste d'IDs ne peut pas être vide.")
    if len(ids) > 500:
        raise HTTPException(status_code=422, detail="Maximum 500 enseignants supprimables en une seule opération.")
    # Empêcher l'auto-suppression
    ids_filtered = [i for i in ids if i != current_user.id]
    from app.models.user import User
    result = await db.execute(
        delete(User).where(User.id.in_(ids_filtered))
    )
    await db.commit()
    return {"deleted": result.rowcount}


@router.delete("/{teacher_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_teacher(teacher_id: uuid.UUID, db: DB, current_user: AdminUser) -> Response:
    await user_service.delete_user(db, teacher_id, current_user.id)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/{teacher_id}/toggle-status", response_model=UserResponse)
async def toggle_status(teacher_id: uuid.UUID, db: DB, _: AdminUser) -> UserResponse:
    return await user_service.toggle_status(db, teacher_id)
