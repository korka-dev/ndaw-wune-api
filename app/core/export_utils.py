"""Utilitaires partagés pour les exports personnalisables (xlsx / csv) du dashboard admin.

Chaque export définit la liste complète de ses colonnes possibles sous la forme
`(clé, en-tête, largeur)`. Le paramètre de requête `fields` (liste de clés séparées
par des virgules) permet à l'utilisateur de ne sélectionner qu'un sous-ensemble de
colonnes. Si `fields` est vide/absent, toutes les colonnes sont exportées
(comportement par défaut, rétro-compatible).
"""
from __future__ import annotations

import csv
import io
from typing import Optional, Sequence

import openpyxl
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill


def parse_fields(fields: Optional[str], all_keys: Sequence[str]) -> list[str]:
    """Retourne la liste des clés à exporter, dans l'ordre d'origine."""
    if not fields:
        return list(all_keys)
    requested = {f.strip() for f in fields.split(",") if f.strip()}
    selected = [k for k in all_keys if k in requested]
    return selected or list(all_keys)


def build_xlsx_response(
    *,
    sheet_title: str,
    columns: Sequence[tuple[str, str, int]],
    rows: Sequence[Sequence],
    fields: Optional[str],
    filename: str,
) -> StreamingResponse:
    """Construit un export xlsx en ne gardant que les colonnes sélectionnées.

    `columns` : liste de tuples (clé, en-tête, largeur) couvrant TOUTES les colonnes possibles.
    `rows` : liste de lignes, chaque ligne ayant autant de valeurs que `columns`.
    """
    all_keys = [c[0] for c in columns]
    selected = parse_fields(fields, all_keys)
    indices = [all_keys.index(k) for k in selected]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="1e6fbf", end_color="1e6fbf", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [columns[i][1] for i in indices]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row[i] for i in indices])

    for col_idx, i in enumerate(indices, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = columns[i][2]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def build_csv_response(
    *,
    columns: Sequence[tuple[str, str]],
    rows: Sequence[Sequence],
    fields: Optional[str],
    filename: str,
) -> StreamingResponse:
    """Construit un export csv en ne gardant que les colonnes sélectionnées.

    `columns` : liste de tuples (clé, en-tête) couvrant TOUTES les colonnes possibles.
    `rows` : liste de lignes, chaque ligne ayant autant de valeurs que `columns`.
    """
    all_keys = [c[0] for c in columns]
    selected = parse_fields(fields, all_keys)
    indices = [all_keys.index(k) for k in selected]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([columns[i][1] for i in indices])
    for row in rows:
        writer.writerow([row[i] for i in indices])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
