#!/usr/bin/env python3
"""
Export des données de terrain — à faire AVANT reset_base_nwv.py
================================================================
Produit un classeur Excel, un onglet par type de donnée, de tout ce que la
réinitialisation va supprimer. Lecture seule : n'écrit rien en base.

Pourquoi : le dump PostgreSQL pris par le reset n'est relisible qu'en
restaurant toute la base, ce qui écraserait le nouvel import. Cet export reste
exploitable tel quel pour la recherche, dans Excel, R ou Stata.

Onglets produits
    Rapports journaliers    rapports_journalier
    Séances                 seances
    Rapports de séance      rapports_prof
    Pointages superviseur   supervisor_presence_checks
    Évaluations élèves      evaluations_eleves
    Remplacements élèves    eleve_remplacements
    Remarques               remarques
    Usage app               usage_logs (agrégé par utilisateur et fonctionnalité)
    Élèves                  eleves (référentiel, pour relire les évaluations)

Les identifiants UUID sont remplacés par les noms (tuteur, superviseur, école,
élève) partout où c'est possible, pour que le fichier reste lisible une fois la
base réinitialisée.

Usage — depuis le dossier backend/
    python3 scripts/export_donnees_terrain.py
    python3 scripts/export_donnees_terrain.py --sortie /tmp/terrain.xlsx

Sur le VPS (le backend tourne en conteneur) :
    docker compose exec -T backend python scripts/export_donnees_terrain.py \\
        --sortie /tmp/terrain.xlsx
    docker compose cp backend:/tmp/terrain.xlsx ./terrain-avant-reset.xlsx
"""

import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Forcer DATABASE_URL vers localhost hors Docker (même trick que les autres scripts) ──
_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv") and (
    "DATABASE_URL" not in os.environ
    or "://db:" in os.environ.get("DATABASE_URL", "")
):
    os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL", _DEFAULT_DB).replace(
        "@db:", "@localhost:"
    )
    if os.environ["DATABASE_URL"].startswith("postgresql+asyncpg://$"):
        os.environ["DATABASE_URL"] = _DEFAULT_DB

import asyncio

from sqlalchemy import text

from app.core.database import AsyncSessionLocal


# Un onglet par requête. L'ordre des colonnes est celui du SELECT.
ONGLETS = [
    ("Rapports journaliers", """
        SELECT r.date_rapport, r.ief, r.commune, r.ecole, r.superviseur, r.nom_tuteur,
               u.name AS compte_tuteur, u.phone AS tel_tuteur,
               r.semaine, r.jour_cours, r.nb_absences, r.absents,
               r.difficultes, r.autres_difficultes, r.description_difficultes,
               r.directeur_venu, r.besoin_appui, r.domaines_appui,
               r.has_observations, r.commentaires, r.reponses_questions,
               r.soumis_en_offline, r.photo_classe_url, r.photos_classe_url,
               r.created_at
        FROM rapports_journalier r
        LEFT JOIN users u ON u.id = r.teacher_id
        ORDER BY r.date_rapport, r.ecole, r.nom_tuteur
    """),
    ("Séances", """
        SELECT s.date_seance, u.name AS tuteur, u.phone AS tel_tuteur,
               ec.name AS ecole, ec.region AS ief, ec.langue,
               s.classe, s.matiere, s.status::text AS statut,
               s.started_at, s.finished_at, s.duree_minutes, s.total_paused_minutes,
               s.nb_eleves_presents, s.nb_eleves_total,
               ps.name AS session, s.created_at
        FROM seances s
        LEFT JOIN users u ON u.id = s.teacher_id
        LEFT JOIN schools ec ON ec.id = u.school_id
        LEFT JOIN program_sessions ps ON ps.id = s.session_id
        ORDER BY s.date_seance
    """),
    ("Rapports de séance", """
        SELECT s.date_seance, u.name AS tuteur, ec.name AS ecole,
               s.classe, s.matiere, s.duree_minutes,
               rp.contenu, rp.difficultes, rp.points_positifs,
               rp.soumis_en_offline, rp.created_at
        FROM rapports_prof rp
        LEFT JOIN seances s ON s.id = rp.seance_id
        LEFT JOIN users u ON u.id = rp.teacher_id
        LEFT JOIN schools ec ON ec.id = u.school_id
        ORDER BY s.date_seance
    """),
    ("Pointages superviseur", """
        SELECT p.date_jour, sup.name AS superviseur, sup.phone AS tel_superviseur,
               tut.name AS tuteur, ec.name AS ecole, ec.region AS ief,
               p.semaine, p.jour_cours, p.present, p.motif, p.created_at
        FROM supervisor_presence_checks p
        LEFT JOIN users sup ON sup.id = p.superviseur_id
        LEFT JOIN users tut ON tut.id = p.teacher_id
        LEFT JOIN schools ec ON ec.id = tut.school_id
        ORDER BY p.date_jour, sup.name
    """),
    ("Évaluations élèves", """
        SELECT ev.date_eval, sup.name AS superviseur,
               e.code_eleve, e.nom AS nom_eleve, e.prenom AS prenom_eleve,
               e.classe, e.genre, e.statut_selection, e.groupe_lecture, e.groupe_maths,
               ec.name AS ecole, ec.region AS ief, ec.langue,
               ev.competence, ev.resultat, ev.commentaire,
               ps.name AS session, ev.created_at
        FROM evaluations_eleves ev
        LEFT JOIN users sup ON sup.id = ev.superviseur_id
        LEFT JOIN eleves e ON e.id = ev.eleve_id
        LEFT JOIN schools ec ON ec.id = e.school_id
        LEFT JOIN program_sessions ps ON ps.id = ev.session_id
        ORDER BY ev.date_eval, ec.name, e.nom
    """),
    ("Remplacements élèves", """
        SELECT rm.date_remplacement, u.name AS tuteur, ec.name AS ecole, rm.classe,
               rm.ancien_eleve_nom, anc.code_eleve AS ancien_code,
               rm.nouveau_eleve_nom, nouv.code_eleve AS nouveau_code,
               rm.motif, rm.created_at
        FROM eleve_remplacements rm
        LEFT JOIN users u ON u.id = rm.teacher_id
        LEFT JOIN schools ec ON ec.id = rm.school_id
        LEFT JOIN eleves anc ON anc.id = rm.ancien_eleve_id
        LEFT JOIN eleves nouv ON nouv.id = rm.nouveau_eleve_id
        ORDER BY rm.date_remplacement
    """),
    ("Remarques", """
        SELECT rq.created_at, rq.user_name, rq.user_role, rq.ecole,
               rq.categorie, rq.message, rq.statut,
               rq.reponse_admin, rq.reponse_admin_at
        FROM remarques rq
        ORDER BY rq.created_at
    """),
    ("Usage app", """
        SELECT user_name, user_role, feature,
               COUNT(*) AS nb_utilisations,
               MIN(created_at)::date AS premier,
               MAX(created_at)::date AS dernier
        FROM usage_logs
        GROUP BY user_name, user_role, feature
        ORDER BY user_name, nb_utilisations DESC
    """),
    ("Élèves", """
        SELECT e.code_eleve, e.nom, e.prenom, e.classe, e.genre, e.statut,
               e.statut_selection, e.groupe_lecture, e.groupe_maths,
               ec.code_ecole, ec.name AS ecole, ec.region AS ief,
               ec.city AS commune, ec.langue
        FROM eleves e
        LEFT JOIN schools ec ON ec.id = e.school_id
        ORDER BY ec.code_ecole, e.classe, e.nom
    """),
]


def cellule(v):
    """Excel n'accepte ni les UUID, ni les listes, ni les timezone-aware."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    if isinstance(v, date):
        return v
    if isinstance(v, (list, tuple)):
        return " | ".join(str(x) for x in v)
    return str(v)


async def main(sortie):
    print(f"Export des données de terrain → {sortie}")
    print("=" * 62)

    wb = Workbook()
    wb.remove(wb.active)
    total = 0

    async with AsyncSessionLocal() as s:
        for titre, requete in ONGLETS:
            res = (await s.execute(text(requete))).mappings().all()
            ws = wb.create_sheet(titre[:31])
            if not res:
                ws.append(["(aucune donnée)"])
                print(f"   {titre:24s}      0")
                continue

            colonnes = list(res[0].keys())
            ws.append(colonnes)
            for c in ws[1]:
                c.font = Font(bold=True)
            for ligne in res:
                ws.append([cellule(ligne[c]) for c in colonnes])

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, nom in enumerate(colonnes, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(38, max(12, len(nom) + 4))

            total += len(res)
            print(f"   {titre:24s} {len(res):6d}")

    wb.save(sortie)
    taille = Path(sortie).stat().st_size
    print("=" * 62)
    print(f"✅  {total} lignes · {taille / 1024 / 1024:.1f} Mo → {sortie}")
    print("   Vérifiez que le fichier s'ouvre AVANT de lancer le reset.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--sortie", default="donnees-terrain.xlsx",
                    help="Chemin du fichier Excel à produire")
    args = ap.parse_args()
    asyncio.run(main(args.sortie))
