#!/usr/bin/env python3
"""
Complète les rattachements tuteur → classe — BaseNWVFinale2026.xlsx
========================================================================
Constat (voir import_groupes_rct_2026.py) : ~1188 élèves du fichier sont
dans une paire (école, classe) qu'aucun tuteur ne couvre dans User.classes.
Le fichier associe pourtant chaque élève à un nom_tuteur/code_tuteur —
dans l'immense majorité des cas ce tuteur existe déjà en base, il lui
manque juste cette classe dans son tableau `classes` (un même tuteur peut
couvrir plusieurs sections dans les petites écoles).

Pour chaque paire (code_ecole, classe_eleve) du fichier non couverte par
un tuteur existant :
  - Un tuteur du même nom (comparaison insensible aux accents/casse)
    existe déjà à cette école → on ajoute la classe à son `classes`
    (append, sans doublon).
  - Sinon → on crée le tuteur (mot de passe temporaire, changement
    obligatoire à la première connexion), rattaché à cette classe.

Idempotent : relancer ne duplique rien (classes dédupliquées, tuteurs
déjà créés retrouvés par nom+école).

Usage :
    python scripts/fix_classes_tuteurs_2026.py <fichier.xlsx>              (dry-run par défaut)
    python scripts/fix_classes_tuteurs_2026.py <fichier.xlsx> --apply       (applique les changements)
"""

import os
import sys
import unicodedata
import uuid
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv") and (
    "DATABASE_URL" not in os.environ
    or "://db:" in os.environ.get("DATABASE_URL", "")
):
    os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL", _DEFAULT_DB).replace(
        "@db:", "@localhost:"
    )
    if os.environ["DATABASE_URL"].startswith("postgresql+asyncpg://$"):
        os.environ["DATABASE_URL"] = _DEFAULT_DB

import asyncio

from sqlalchemy import text

from app.core.database import AsyncSessionLocal
from app.core.redis import invalidate_sync_caches
from app.core.security import hash_password

DRY_RUN = "--apply" not in sys.argv
FILE_ARG = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
DEFAULT_PASSWORD = "Passer123"


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn"
    ).strip().upper()


def read_pairs(xlsx_path: str) -> dict[tuple[int, str], tuple[str, str]]:
    """(code_ecole, classe_eleve) -> (code_tuteur, nom_tuteur), une entrée par paire."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]
    required = {"code_ecole", "classe_eleve", "code_tuteur", "nom_tuteur"}
    missing = required - set(headers)
    if missing:
        sys.exit(f"❌  Colonnes manquantes dans le fichier : {missing}")

    pairs: dict[tuple[int, str], tuple[str, str]] = {}
    for values in rows_iter:
        row = dict(zip(headers, values))
        if row.get("code_ecole") is None or row.get("classe_eleve") is None:
            continue
        key = (int(row["code_ecole"]), str(row["classe_eleve"]).strip())
        pairs.setdefault(key, (row.get("code_tuteur"), str(row.get("nom_tuteur") or "").strip()))
    return pairs


async def run(xlsx_path: str):
    mode = "DRY-RUN (aucune modification)" if DRY_RUN else "APPLICATION RÉELLE"
    print(f"\n{'='*60}")
    print(f"  Rattachement tuteur -> classe 2026 — {mode}")
    print(f"{'='*60}\n")

    print("📂  Lecture du fichier Excel…")
    pairs = read_pairs(xlsx_path)
    print(f"   {len(pairs)} paires (école, classe) distinctes dans le fichier")

    async with AsyncSessionLocal() as s:
        school_rows = (await s.execute(text(
            "SELECT id, code_ecole, name FROM schools WHERE code_ecole IS NOT NULL"
        ))).fetchall()
        school_by_code = {r[1]: (r[0], r[2]) for r in school_rows}

        teacher_rows = (await s.execute(text(
            "SELECT id, school_id, name, classes FROM users WHERE role = 'enseignant'"
        ))).fetchall()
        # (school_id) -> liste de (id, name_normalise, classes_actuelles)
        teachers_by_school: dict[str, list] = {}
        for tid, school_id, name, classes in teacher_rows:
            teachers_by_school.setdefault(str(school_id), []).append(
                {"id": tid, "name": name, "norm": strip_accents(name), "classes": list(classes or [])}
            )

        covered = {(str(t["id"]), c) for lst in teachers_by_school.values() for t in lst for c in t["classes"]}
        # pour vérif rapide "cette (école,classe) est-elle déjà couverte par un tuteur"
        school_classe_covered = set()
        for school_id, lst in teachers_by_school.items():
            for t in lst:
                for c in t["classes"]:
                    school_classe_covered.add((school_id, c))

        updated = 0
        created = 0
        already_ok = 0
        errors: list[str] = []

        print("\n⏳  Traitement…")
        for (code_ecole, classe), (code_tuteur, nom_tuteur) in pairs.items():
            school = school_by_code.get(code_ecole)
            if school is None:
                errors.append(f"code_ecole={code_ecole} introuvable en base — ignoré.")
                continue
            school_id, school_name = school
            key = (str(school_id), classe)

            if key in school_classe_covered:
                already_ok += 1
                continue

            if not nom_tuteur:
                errors.append(f"École {school_name} / classe '{classe}' : nom_tuteur vide dans le fichier — ignoré.")
                continue

            nom_norm = strip_accents(nom_tuteur)
            candidates = [t for t in teachers_by_school.get(str(school_id), []) if t["norm"] == nom_norm]

            if len(candidates) == 1:
                teacher = candidates[0]
                if classe not in teacher["classes"]:
                    teacher["classes"].append(classe)
                    school_classe_covered.add(key)
                    if not DRY_RUN:
                        await s.execute(text(
                            "UPDATE users SET classes = :classes, updated_at = NOW() WHERE id = :id"
                        ), {"classes": teacher["classes"], "id": teacher["id"]})
                    updated += 1
                    print(f"   + {nom_tuteur} ({school_name}) : ajout classe '{classe}' -> {teacher['classes']}")
                continue

            if len(candidates) > 1:
                errors.append(
                    f"École {school_name} / classe '{classe}' : plusieurs tuteurs nommés "
                    f"'{nom_tuteur}' — ignoré (ambigu)."
                )
                continue

            # Aucun tuteur de ce nom à cette école → création
            new_id = uuid.uuid4()
            print(f"   ★ Création tuteur {nom_tuteur} ({school_name}) — classe '{classe}'")
            if not DRY_RUN:
                await s.execute(text(
                    """
                    INSERT INTO users
                        (id, name, password_hash, role, status, app_access, must_change_password,
                         school_id, classes, created_at, updated_at)
                    VALUES
                        (:id, :name, :password_hash, 'enseignant', 'actif', 'full', true,
                         :school_id, :classes, NOW(), NOW())
                    """
                ), {
                    "id": str(new_id), "name": nom_tuteur, "password_hash": hash_password(DEFAULT_PASSWORD),
                    "school_id": str(school_id), "classes": [classe],
                })
            teachers_by_school.setdefault(str(school_id), []).append(
                {"id": new_id, "name": nom_tuteur, "norm": nom_norm, "classes": [classe]}
            )
            school_classe_covered.add(key)
            created += 1

        if not DRY_RUN:
            await s.commit()
            print("\n   Commit effectué ✓")
            await invalidate_sync_caches()
        else:
            print("\n   [DRY-RUN] Aucune écriture effectuée")

        print(f"\n{'='*60}")
        print("📊  RÉSUMÉ")
        print(f"{'='*60}")
        print(f"  Paires (école, classe) dans le fichier : {len(pairs)}")
        print(f"  Déjà couvertes                          : {already_ok}")
        print(f"  Tuteurs mis à jour (classe ajoutée)     : {updated}")
        print(f"  Tuteurs créés                           : {created}")
        print(f"  Erreurs / avertissements                : {len(errors)}")
        for e in errors[:30]:
            print(f"    ⚠ {e}")
        if len(errors) > 30:
            print(f"    … et {len(errors) - 30} autres")
        print(f"{'='*60}\n")
        if created and not DRY_RUN:
            print(f"  Mot de passe temporaire des nouveaux tuteurs : {DEFAULT_PASSWORD}")
            print("  (changement obligatoire à la première connexion — pensez à")
            print("   renseigner leur téléphone/email dans l'admin pour la connexion)\n")
        if DRY_RUN:
            print("  → Relancer avec --apply pour exécuter réellement\n")


if __name__ == "__main__":
    if not FILE_ARG:
        sys.exit("Usage : python scripts/fix_classes_tuteurs_2026.py <fichier.xlsx> [--apply]")
    asyncio.run(run(FILE_ARG))
