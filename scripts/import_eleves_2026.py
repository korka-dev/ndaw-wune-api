#!/usr/bin/env python3
"""
Remplacement de la liste d'élèves — BaseNWVFinale2026.xlsx
============================================================
Actions :
  1. Supprime TOUS les élèves existants (cascade sur evaluation_tirages / evaluations_eleves)
  2. Importe les élèves du fichier Excel, rattachés à leur école (code_ecole)
     et à la session de programme active

Colonnes ignorées volontairement (pas de colonne équivalente sur Eleve) :
  superviseur, tuteur, groupe_lecture, groupe_maths, statut_selection

Usage :
    python scripts/import_eleves_2026.py <chemin_vers_xlsx>              (dry-run par défaut)
    python scripts/import_eleves_2026.py <chemin_vers_xlsx> --apply       (applique les changements)
"""

import os
import sys
import uuid
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Forcer DATABASE_URL vers localhost hors Docker (même trick que cleanup_eleves.py) ──
_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv") and (
    "DATABASE_URL" not in os.environ
    or "://db:" in os.environ.get("DATABASE_URL", "")
):
    os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL", _DEFAULT_DB).replace(
        "@db:", "@localhost:"
    )
    if os.environ["DATABASE_URL"].startswith("postgresql+asyncpg://$"):
        os.environ["DATABASE_URL"] = _DEFAULT_DB

import asyncio

from sqlalchemy import text

from app.core.database import AsyncSessionLocal

DRY_RUN = "--apply" not in sys.argv
FILE_ARG = next((a for a in sys.argv[1:] if not a.startswith("--")), None)


def normalize_genre(sexe_str) -> str | None:
    s = str(sexe_str or "").strip().lower()
    if s in ("fille", "f"):
        return "F"
    if s in ("garcon", "garçon", "g", "m"):
        return "M"
    return str(sexe_str).strip() or None


def split_name(full_name: str) -> tuple[str | None, str]:
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return (None, parts[0])
    return (" ".join(parts[:-1]), parts[-1])


def read_rows(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]
    required = {"code_ecole", "classe_eleve", "nom_eleve", "sexe_str", "code_eleve"}
    missing = required - set(headers)
    if missing:
        sys.exit(f"❌  Colonnes manquantes dans le fichier : {missing}")

    rows = []
    for values in rows_iter:
        row = dict(zip(headers, values))
        if row.get("nom_eleve") is None or row.get("code_ecole") is None or row.get("classe_eleve") is None:
            continue
        rows.append(row)
    return rows


async def run(xlsx_path: str):
    mode = "DRY-RUN (aucune modification)" if DRY_RUN else "APPLICATION RÉELLE"
    print(f"\n{'='*60}")
    print(f"  Remplacement élèves 2026 — {mode}")
    print(f"{'='*60}\n")

    print("📂  Lecture du fichier Excel…")
    rows = read_rows(xlsx_path)
    print(f"   {len(rows)} lignes valides à traiter")

    async with AsyncSessionLocal() as s:
        session_row = (await s.execute(text(
            "SELECT id, name FROM program_sessions WHERE status = 'active' "
            "ORDER BY created_at DESC LIMIT 1"
        ))).fetchone()
        if not session_row:
            sys.exit("❌  Aucune session active trouvée dans program_sessions.")
        session_id, session_name = session_row
        print(f"🔎  Session active : « {session_name} » (id={session_id})")

        school_rows = (await s.execute(text(
            "SELECT id, code_ecole FROM schools WHERE code_ecole IS NOT NULL"
        ))).fetchall()
        school_by_code = {r[1]: r[0] for r in school_rows}
        print(f"🏫  {len(school_by_code)} écoles avec code_ecole en base")

        file_codes = {int(r["code_ecole"]) for r in rows}
        missing_codes = sorted(file_codes - set(school_by_code.keys()))
        if missing_codes:
            sys.exit(f"❌  code_ecole absents de la base, import annulé : {missing_codes}")

        total_before = (await s.execute(text("SELECT COUNT(*) FROM eleves"))).scalar()
        print(f"🗑️   {total_before} élèves actuellement en base (seront supprimés)")

        if not DRY_RUN:
            result = await s.execute(text("DELETE FROM eleves"))
            print(f"   ✓ {result.rowcount} élèves supprimés")

        print("\n⏳  Insertion des nouveaux élèves…")
        inserted = 0
        skipped = 0
        for row in rows:
            code_ecole = int(row["code_ecole"])
            school_id = school_by_code[code_ecole]
            classe = str(row["classe_eleve"]).strip()
            full_name = str(row["nom_eleve"]).strip()
            prenom, nom = split_name(full_name)
            genre = normalize_genre(row["sexe_str"])
            code_eleve = str(row["code_eleve"]).strip() if row.get("code_eleve") is not None else None

            if not DRY_RUN:
                result = await s.execute(text(
                    """
                    INSERT INTO eleves
                        (id, nom, prenom, classe, genre, statut, code_eleve, school_id, session_id, created_at, updated_at)
                    VALUES
                        (:id, :nom, :prenom, :classe, :genre, 'actif', :code_eleve, :school_id, :session_id, NOW(), NOW())
                    ON CONFLICT ON CONSTRAINT uq_eleve_school_classe_nom DO NOTHING
                    """
                ), {
                    "id": str(uuid.uuid4()), "nom": nom, "prenom": prenom, "classe": classe,
                    "genre": genre, "code_eleve": code_eleve, "school_id": str(school_id),
                    "session_id": str(session_id),
                })
                if result.rowcount == 1:
                    inserted += 1
                else:
                    skipped += 1
            else:
                inserted += 1

        if not DRY_RUN:
            await s.commit()
            print("   Commit effectué ✓")
        else:
            print("   [DRY-RUN] Aucune écriture effectuée")

        print(f"\n{'='*60}")
        print("📊  RÉSUMÉ")
        print(f"{'='*60}")
        print(f"  Élèves existants avant  : {total_before}")
        print(f"  Lignes traitées         : {len(rows)}")
        print(f"  Élèves insérés          : {inserted}")
        print(f"  Doublons ignorés        : {skipped}")
        print(f"{'='*60}\n")
        if DRY_RUN:
            print("  → Relancer avec --apply pour exécuter réellement\n")


if __name__ == "__main__":
    if not FILE_ARG:
        sys.exit("Usage : python scripts/import_eleves_2026.py <fichier.xlsx> [--apply]")
    asyncio.run(run(FILE_ARG))
