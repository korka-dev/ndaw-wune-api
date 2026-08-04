#!/usr/bin/env python3
"""
Import groupes RCT + statut de sélection — BaseNWVFinale2026.xlsx
====================================================================
Complète import_eleves_2026.py : renseigne sur Eleve les trois champs
ignorés par ce dernier — statut_selection, groupe_lecture, groupe_maths —
sans jamais supprimer ni recréer d'élève.

Comportement (upsert, jamais de DELETE) :
  - Élève déjà en base (match par code_eleve) → met à jour uniquement
    statut_selection / groupe_lecture / groupe_maths.
  - Élève absent de la base → le crée (nom/prénom/classe/genre/école),
    avec ces trois champs renseignés dès la création.

Un DELETE sur eleves supprimerait en cascade evaluations_eleves et
evaluation_tirages (ON DELETE CASCADE) — c'est précisément ce que ce
script évite.

Usage :
    python scripts/import_groupes_rct_2026.py <fichier.xlsx>              (dry-run par défaut)
    python scripts/import_groupes_rct_2026.py <fichier.xlsx> --apply       (applique les changements)
"""

import os
import sys
import uuid
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv") and (
    "DATABASE_URL" not in os.environ
    or "://db:" in os.environ.get("DATABASE_URL", "")
):
    os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL", _DEFAULT_DB).replace(
        "@db:", "@localhost:"
    )
    if os.environ["DATABASE_URL"].startswith("postgresql+asyncpg://$"):
        os.environ["DATABASE_URL"] = _DEFAULT_DB

import asyncio

from sqlalchemy import text

from app.core.database import AsyncSessionLocal

DRY_RUN = "--apply" not in sys.argv
FILE_ARG = next((a for a in sys.argv[1:] if not a.startswith("--")), None)


def normalize_genre(sexe_str) -> str | None:
    s = str(sexe_str or "").strip().lower()
    if s in ("fille", "f"):
        return "F"
    if s in ("garcon", "garçon", "g", "m"):
        return "M"
    return str(sexe_str).strip() or None


def split_name(full_name: str) -> tuple[str | None, str]:
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return (None, parts[0])
    return (" ".join(parts[:-1]), parts[-1])


def read_rows(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]
    required = {"code_ecole", "classe_eleve", "nom_eleve", "sexe_str", "code_eleve",
                "statut_selection", "groupe_lecture", "groupe_maths"}
    missing = required - set(headers)
    if missing:
        sys.exit(f"❌  Colonnes manquantes dans le fichier : {missing}")

    rows = []
    for values in rows_iter:
        row = dict(zip(headers, values))
        if row.get("nom_eleve") is None or row.get("code_ecole") is None or row.get("code_eleve") is None:
            continue
        rows.append(row)
    return rows


async def run(xlsx_path: str):
    mode = "DRY-RUN (aucune modification)" if DRY_RUN else "APPLICATION RÉELLE"
    print(f"\n{'='*60}")
    print(f"  Groupes RCT + statut sélection 2026 — {mode}")
    print(f"{'='*60}\n")

    print("📂  Lecture du fichier Excel…")
    rows = read_rows(xlsx_path)
    print(f"   {len(rows)} lignes valides à traiter")

    async with AsyncSessionLocal() as s:
        session_row = (await s.execute(text(
            "SELECT id, name FROM program_sessions WHERE status = 'active' "
            "ORDER BY created_at DESC LIMIT 1"
        ))).fetchone()
        if not session_row:
            sys.exit("❌  Aucune session active trouvée dans program_sessions.")
        session_id, session_name = session_row
        print(f"🔎  Session active : « {session_name} » (id={session_id})")

        school_rows = (await s.execute(text(
            "SELECT id, code_ecole FROM schools WHERE code_ecole IS NOT NULL"
        ))).fetchall()
        school_by_code = {r[1]: r[0] for r in school_rows}
        print(f"🏫  {len(school_by_code)} écoles avec code_ecole en base")

        file_codes_ecole = {int(r["code_ecole"]) for r in rows}
        missing_codes = sorted(file_codes_ecole - set(school_by_code.keys()))
        if missing_codes:
            sys.exit(f"❌  code_ecole absents de la base, import annulé : {missing_codes}")

        existing_rows = (await s.execute(text(
            "SELECT code_eleve, id FROM eleves WHERE code_eleve IS NOT NULL"
        ))).fetchall()
        eleve_by_code = {r[0]: r[1] for r in existing_rows}
        print(f"👤  {len(eleve_by_code)} élèves déjà en base (code_eleve renseigné)")

        # Repli — élève déjà en base mais sous un autre code_eleve (renumérotation
        # côté source) : identifié par (school_id, classe, nom, prenom).
        name_rows = (await s.execute(text(
            "SELECT school_id, classe, nom, prenom, id, code_eleve FROM eleves"
        ))).fetchall()
        eleve_by_identity: dict[tuple, list] = {}
        for r in name_rows:
            key = (str(r[0]), r[1], r[2], r[3])
            eleve_by_identity.setdefault(key, []).append((r[4], r[5]))

        updated = 0
        recoded = 0
        created = 0
        skipped_dup = 0
        errors: list[str] = []

        print("\n⏳  Traitement…")
        for i, row in enumerate(rows):
            code_eleve = str(row["code_eleve"]).strip()
            statut_selection = str(row["statut_selection"]).strip() if row.get("statut_selection") else None
            groupe_lecture = str(row["groupe_lecture"]).strip() if row.get("groupe_lecture") else None
            groupe_maths = str(row["groupe_maths"]).strip() if row.get("groupe_maths") else None

            if code_eleve in eleve_by_code:
                if not DRY_RUN:
                    await s.execute(text(
                        """
                        UPDATE eleves
                        SET statut_selection = :statut_selection,
                            groupe_lecture   = :groupe_lecture,
                            groupe_maths     = :groupe_maths,
                            updated_at       = NOW()
                        WHERE code_eleve = :code_eleve
                        """
                    ), {
                        "statut_selection": statut_selection,
                        "groupe_lecture": groupe_lecture,
                        "groupe_maths": groupe_maths,
                        "code_eleve": code_eleve,
                    })
                updated += 1
                continue

            code_ecole = int(row["code_ecole"])
            school_id = school_by_code[code_ecole]
            classe = str(row["classe_eleve"]).strip()
            full_name = str(row["nom_eleve"]).strip()
            prenom, nom = split_name(full_name)
            genre = normalize_genre(row["sexe_str"])

            identity_key = (str(school_id), classe, nom, prenom)
            candidates = eleve_by_identity.get(identity_key, [])
            if len(candidates) == 1:
                # Même élève, code_eleve renuméroté côté source — on aligne.
                existing_id, existing_code = candidates[0]
                if not DRY_RUN:
                    await s.execute(text(
                        """
                        UPDATE eleves
                        SET code_eleve       = :code_eleve,
                            statut_selection = :statut_selection,
                            groupe_lecture   = :groupe_lecture,
                            groupe_maths     = :groupe_maths,
                            updated_at       = NOW()
                        WHERE id = :id
                        """
                    ), {
                        "code_eleve": code_eleve,
                        "statut_selection": statut_selection,
                        "groupe_lecture": groupe_lecture,
                        "groupe_maths": groupe_maths,
                        "id": existing_id,
                    })
                recoded += 1
                if existing_code != code_eleve:
                    errors.append(
                        f"Ligne {i + 2} ({full_name}, {classe}) : code_eleve renuméroté "
                        f"{existing_code} → {code_eleve}"
                    )
                continue
            elif len(candidates) > 1:
                errors.append(
                    f"Ligne {i + 2} ({full_name}, {classe}) : plusieurs élèves homonymes "
                    f"dans la même école/classe — ignoré (ambigu)."
                )
                continue

            # Élève réellement absent → création
            if not DRY_RUN:
                result = await s.execute(text(
                    """
                    INSERT INTO eleves
                        (id, nom, prenom, classe, genre, statut, code_eleve, school_id, session_id,
                         statut_selection, groupe_lecture, groupe_maths, created_at, updated_at)
                    VALUES
                        (:id, :nom, :prenom, :classe, :genre, 'actif', :code_eleve, :school_id, :session_id,
                         :statut_selection, :groupe_lecture, :groupe_maths, NOW(), NOW())
                    ON CONFLICT ON CONSTRAINT uq_eleve_school_classe_nom DO NOTHING
                    """
                ), {
                    "id": str(uuid.uuid4()), "nom": nom, "prenom": prenom, "classe": classe,
                    "genre": genre, "code_eleve": code_eleve, "school_id": str(school_id),
                    "session_id": str(session_id), "statut_selection": statut_selection,
                    "groupe_lecture": groupe_lecture, "groupe_maths": groupe_maths,
                })
                if result.rowcount == 1:
                    created += 1
                else:
                    skipped_dup += 1
            else:
                created += 1

        if not DRY_RUN:
            await s.commit()
            print("   Commit effectué ✓")
        else:
            print("   [DRY-RUN] Aucune écriture effectuée")

        print(f"\n{'='*60}")
        print("📊  RÉSUMÉ")
        print(f"{'='*60}")
        print(f"  Lignes traitées         : {len(rows)}")
        print(f"  Élèves mis à jour       : {updated}")
        print(f"  Élèves renumérotés      : {recoded}")
        print(f"  Élèves créés            : {created}")
        print(f"  Doublons ignorés        : {skipped_dup}")
        print(f"  Erreurs / avertissements: {len(errors)}")
        for e in errors[:30]:
            print(f"    ⚠ {e}")
        if len(errors) > 30:
            print(f"    … et {len(errors) - 30} autres")
        print(f"{'='*60}\n")
        if DRY_RUN:
            print("  → Relancer avec --apply pour exécuter réellement\n")


if __name__ == "__main__":
    if not FILE_ARG:
        sys.exit("Usage : python scripts/import_groupes_rct_2026.py <fichier.xlsx> [--apply]")
    asyncio.run(run(FILE_ARG))
