#!/usr/bin/env python3
"""
Import NdawWune.xlsx → Base de données NDAW WUNE
=================================================
Script complet d'import (openpyxl uniquement — pas de pandas).
Fonctionne en Docker et en local.

Insère dans l'ordre :
  1. Schools       (écoles avec code_ecole, region=IEF, city=commune, director)
  2. SchoolClasse  (classes par école, avec effectif calculé)
  3. Élèves        (avec code_eleve, dédupliqués)
  4. Superviseurs  (directeurs → Users role=superviseur, dédupliqués)
  5. Liaisons      (chaque superviseur reçoit les UUIDs des enseignants de son école)

Colonnes attendues dans l'Excel :
  code_ief, ief, code_commune, commune, Directeur, code_ecole, ecole,
  classe_num, code_eleve, nom_eleve, code_Classe, selection

Usage (Docker — production) :
    # 1. Copier le fichier dans le container
    docker cp NdawWune.xlsx backend:/tmp/NdawWune.xlsx

    # 2. Lancer le script
    docker exec backend python scripts/import_ndawwune.py /tmp/NdawWune.xlsx

    # Mode dry-run (voir sans toucher la DB)
    docker exec backend python scripts/import_ndawwune.py /tmp/NdawWune.xlsx --dry-run

Usage (local) :
    cd backend
    .venv/bin/python scripts/import_ndawwune.py /chemin/vers/NdawWune.xlsx
"""

import asyncio
import os
import sys
import uuid
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

# ── Rendre "app" importable ──────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── DATABASE_URL — fonctionne en Docker ET en local ─────────────────────────
_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv"):
    raw = os.environ.get("DATABASE_URL", _DEFAULT_DB)
    if "://db:" in raw:
        raw = raw.replace("@db:", "@localhost:")
    if raw.startswith("postgresql+asyncpg://$"):
        raw = _DEFAULT_DB
    os.environ["DATABASE_URL"] = raw

from app.core.database import AsyncSessionLocal
from app.core.security import hash_password
from app.models.eleve import Eleve
from app.models.school import School
from app.models.school_classe import SchoolClasse
from app.models.user import User, UserRole, UserStatus

SUPERVISOR_DEFAULT_PASSWORD = "NdawWune2025!"


# ── Lecture Excel avec openpyxl ──────────────────────────────────────────────

def read_excel(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter)
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(raw_headers)]

    data = []
    for row in rows_iter:
        record = {}
        for h, val in zip(headers, row):
            record[h] = val
        # Ignorer les lignes vides
        if not record.get("ecole") or not record.get("nom_eleve"):
            continue
        # Nettoyer les chaînes
        for k in ("ief", "commune", "Directeur", "ecole", "nom_eleve", "code_Classe", "selection"):
            if k in record and record[k] is not None:
                record[k] = str(record[k]).strip()
        data.append(record)

    wb.close()
    return data


# ── Helpers ──────────────────────────────────────────────────────────────────

def split_nom_prenom(full_name: str) -> tuple[str, str | None]:
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return parts[0], None
    return parts[-1], " ".join(parts[:-1])


def extract_niveau(code_classe: str) -> str:
    c = code_classe.strip().upper()
    for prefix in ("CE1", "CE2", "CP", "CM1", "CM2"):
        if c.startswith(prefix):
            return prefix
    return c


# ── Étape 1 — Écoles ────────────────────────────────────────────────────────

async def import_schools(session, rows: list[dict]) -> dict[int, uuid.UUID]:
    seen_codes: set[int] = set()
    ecoles: list[dict] = []
    for r in rows:
        code = int(r["code_ecole"])
        if code in seen_codes:
            continue
        seen_codes.add(code)
        ecoles.append(r)

    existing_res = await session.execute(select(School))
    existing_by_code: dict[int, School] = {}
    existing_by_name: dict[str, School] = {}
    for s in existing_res.scalars():
        if s.code_ecole is not None:
            existing_by_code[s.code_ecole] = s
        existing_by_name[s.name.strip().upper()] = s

    school_map: dict[int, uuid.UUID] = {}
    created = 0
    updated = 0

    for r in ecoles:
        code = int(r["code_ecole"])
        name = str(r["ecole"]).strip()
        ief = str(r["ief"]).strip() if r.get("ief") else None
        commune = str(r["commune"]).strip() if r.get("commune") else None
        directeur = str(r["Directeur"]).strip() if r.get("Directeur") else None

        if code in existing_by_code:
            school = existing_by_code[code]
            changed = False
            if school.director != directeur:
                school.director = directeur
                changed = True
            if school.region != ief:
                school.region = ief
                changed = True
            if school.city != commune:
                school.city = commune
                changed = True
            if changed:
                updated += 1
            school_map[code] = school.id
            continue

        if name.upper() in existing_by_name:
            school = existing_by_name[name.upper()]
            school.code_ecole = code
            school.director = directeur
            school.region = ief
            school.city = commune
            school_map[code] = school.id
            updated += 1
            continue

        school = School(
            id=uuid.uuid4(),
            name=name,
            code_ecole=code,
            region=ief,
            city=commune,
            director=directeur,
        )
        session.add(school)
        school_map[code] = school.id
        created += 1

    await session.flush()
    total_existing = len(school_map) - created
    print(f"  ✓ Écoles : {created} créées, {updated} mises à jour, {total_existing - updated} inchangées → {len(school_map)} total")
    return school_map


# ── Étape 2 — Classes ───────────────────────────────────────────────────────

async def import_classes(
    session, rows: list[dict], school_map: dict[int, uuid.UUID]
) -> dict[tuple[int, str], uuid.UUID]:
    # Compter les effectifs et lister les classes uniques
    effectifs: dict[tuple[int, str], int] = defaultdict(int)
    unique_classes: dict[tuple[int, str], str] = {}
    for r in rows:
        code_ecole = int(r["code_ecole"])
        classe_name = str(r["code_Classe"]).strip()
        key = (code_ecole, classe_name)
        effectifs[key] += 1
        if key not in unique_classes:
            unique_classes[key] = classe_name

    existing_res = await session.execute(select(SchoolClasse))
    school_id_to_code: dict[uuid.UUID, int] = {v: k for k, v in school_map.items()}
    existing_map: dict[tuple[int, str], uuid.UUID] = {}
    for c in existing_res.scalars():
        code = school_id_to_code.get(c.school_id)
        if code is not None:
            existing_map[(code, c.name)] = c.id

    classe_map: dict[tuple[int, str], uuid.UUID] = dict(existing_map)
    created = 0

    for key, classe_name in sorted(unique_classes.items()):
        if key in classe_map:
            continue

        code_ecole = key[0]
        school_id = school_map.get(code_ecole)
        if not school_id:
            continue

        sc = SchoolClasse(
            id=uuid.uuid4(),
            name=classe_name,
            niveau=extract_niveau(classe_name),
            effectif=effectifs[key],
            school_id=school_id,
        )
        session.add(sc)
        classe_map[key] = sc.id
        created += 1

    await session.flush()
    print(f"  ✓ Classes : {created} créées, {len(existing_map)} déjà existantes → {len(classe_map)} total")
    return classe_map


# ── Étape 3 — Élèves ────────────────────────────────────────────────────────

async def import_eleves(
    session, rows: list[dict], school_map: dict[int, uuid.UUID]
) -> int:
    existing_res = await session.execute(select(Eleve))
    existing_set: set[tuple] = set()
    existing_codes: set[str] = set()
    for e in existing_res.scalars():
        existing_set.add((str(e.school_id), e.classe, e.nom, e.prenom or ""))
        if e.code_eleve:
            existing_codes.add(e.code_eleve)

    created = 0
    skipped = 0
    errors = 0

    for r in rows:
        code_ecole = int(r["code_ecole"])
        classe = str(r["code_Classe"]).strip()
        full_name = str(r["nom_eleve"]).strip() if r.get("nom_eleve") else ""
        code_eleve = str(int(r["code_eleve"])) if r.get("code_eleve") is not None else None

        if not full_name:
            errors += 1
            continue

        school_id = school_map.get(code_ecole)
        if not school_id:
            errors += 1
            continue

        nom, prenom = split_nom_prenom(full_name)
        key = (str(school_id), classe, nom, prenom or "")

        if key in existing_set:
            skipped += 1
            continue

        if code_eleve and code_eleve in existing_codes:
            skipped += 1
            continue

        eleve = Eleve(
            id=uuid.uuid4(),
            nom=nom,
            prenom=prenom,
            code_eleve=code_eleve,
            classe=classe,
            genre=None,
            statut="actif",
            school_id=school_id,
            session_id=None,
        )
        session.add(eleve)
        existing_set.add(key)
        if code_eleve:
            existing_codes.add(code_eleve)
        created += 1

        if created % 500 == 0:
            await session.flush()
            print(f"    … {created} élèves insérés")

    await session.flush()
    print(f"  ✓ Élèves : {created} créés, {skipped} doublons ignorés, {errors} erreurs → {created + skipped} traités")
    return created


# ── Étape 4 — Superviseurs (Directeurs) ─────────────────────────────────────

async def import_superviseurs(
    session, rows: list[dict], school_map: dict[int, uuid.UUID]
) -> dict[int, uuid.UUID]:
    seen_codes: set[int] = set()
    directeurs: list[dict] = []
    for r in rows:
        code = int(r["code_ecole"])
        if code in seen_codes:
            continue
        seen_codes.add(code)
        directeurs.append(r)

    existing_res = await session.execute(
        select(User).where(User.role == UserRole.superviseur)
    )
    existing_sups: dict[uuid.UUID, User] = {}
    for u in existing_res.scalars():
        if u.school_id:
            existing_sups[u.school_id] = u

    pwd_hash = hash_password(SUPERVISOR_DEFAULT_PASSWORD)
    sup_map: dict[int, uuid.UUID] = {}
    created = 0
    skipped = 0

    for r in directeurs:
        code_ecole = int(r["code_ecole"])
        directeur_name = str(r["Directeur"]).strip() if r.get("Directeur") else None
        if not directeur_name:
            continue

        school_id = school_map.get(code_ecole)
        if not school_id:
            continue

        if school_id in existing_sups:
            sup_map[code_ecole] = existing_sups[school_id].id
            skipped += 1
            continue

        sup = User(
            id=uuid.uuid4(),
            name=directeur_name,
            email=None,
            phone=None,
            password_hash=pwd_hash,
            role=UserRole.superviseur,
            status=UserStatus.actif,
            must_change_password=True,
            school_id=school_id,
            classes=[],
            niveau=None,
        )
        session.add(sup)
        sup_map[code_ecole] = sup.id
        created += 1

    await session.flush()
    print(f"  ✓ Superviseurs : {created} créés, {skipped} déjà existants → {len(sup_map)} total")
    return sup_map


# ── Étape 5 — Liaison superviseurs ↔ enseignants ────────────────────────────

async def link_superviseurs_to_teachers(
    session, school_map: dict[int, uuid.UUID], sup_map: dict[int, uuid.UUID]
) -> None:
    teachers_res = await session.execute(
        select(User).where(User.role == UserRole.enseignant)
    )
    teachers_by_school: dict[uuid.UUID, list[str]] = defaultdict(list)
    for t in teachers_res.scalars():
        if t.school_id:
            teachers_by_school[t.school_id].append(str(t.id))

    sup_ids = list(sup_map.values())
    if not sup_ids:
        print("  ⚠ Aucun superviseur à lier")
        return

    sups_res = await session.execute(
        select(User).where(User.id.in_(sup_ids))
    )
    linked = 0
    no_teachers = 0

    for sup in sups_res.scalars():
        teacher_ids = teachers_by_school.get(sup.school_id, [])

        if not teacher_ids:
            no_teachers += 1
            continue

        existing_ids = set(sup.classes or [])
        new_ids = set(teacher_ids)
        merged = sorted(existing_ids | new_ids)

        if set(sup.classes or []) != set(merged):
            sup.classes = merged
            linked += 1

    await session.flush()
    print(f"  ✓ Liaisons : {linked} superviseurs liés à leurs enseignants, {no_teachers} sans enseignant dans leur école")


# ── Résumé ───────────────────────────────────────────────────────────────────

def print_summary(rows: list[dict]) -> None:
    iefs = set()
    communes = set()
    ecoles = set()
    directeurs = set()
    classes = set()
    codes_eleves = set()
    titulaires = 0
    remplacants = 0
    ief_counts: dict[str, int] = defaultdict(int)

    for r in rows:
        iefs.add(r.get("ief"))
        communes.add(r.get("commune"))
        ecoles.add(r.get("ecole"))
        directeurs.add(r.get("Directeur"))
        classes.add(r.get("code_Classe"))
        codes_eleves.add(r.get("code_eleve"))
        sel = r.get("selection", "")
        if sel == "Titulaire":
            titulaires += 1
        elif sel == "Remplacant":
            remplacants += 1
        ief_name = r.get("ief", "?")
        ief_counts[ief_name] += 1

    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║              RÉSUMÉ DES DONNÉES EXCEL                   ║")
    print("╠══════════════════════════════════════════════════════════╣")
    print(f"║  Lignes totales        : {len(rows):>6}                        ║")
    print(f"║  IEFs                  : {len(iefs):>6}                        ║")
    print(f"║  Communes              : {len(communes):>6}                        ║")
    print(f"║  Écoles                : {len(ecoles):>6}                        ║")
    print(f"║  Directeurs (→ sup.)   : {len(directeurs):>6}                        ║")
    print(f"║  Classes               : {len(classes):>6}                        ║")
    print(f"║  Élèves                : {len(codes_eleves):>6}                        ║")
    print(f"║  Titulaires            : {titulaires:>6}                        ║")
    print(f"║  Remplaçants           : {remplacants:>6}                        ║")
    print("╚══════════════════════════════════════════════════════════╝")

    print("\n  Répartition par IEF :")
    for ief_name in sorted(ief_counts.keys()):
        print(f"    • {ief_name:<25} {ief_counts[ief_name]:>5} élèves")
    print()


# ── Main ─────────────────────────────────────────────────────────────────────

async def run(xlsx_path: str, dry_run: bool = False) -> None:
    print(f"\n📂  Lecture de {xlsx_path} …")
    rows = read_excel(xlsx_path)
    print(f"  {len(rows)} lignes valides")
    print_summary(rows)

    if dry_run:
        print("🔒  Mode dry-run — aucune modification en base.")
        print("    Relancez sans --dry-run pour exécuter l'import.")
        return

    async with AsyncSessionLocal() as session:
        try:
            print("── [1/5] Écoles ─────────────────────────────────────────")
            school_map = await import_schools(session, rows)

            print("── [2/5] Classes ────────────────────────────────────────")
            await import_classes(session, rows, school_map)

            print("── [3/5] Élèves ─────────────────────────────────────────")
            await import_eleves(session, rows, school_map)

            print("── [4/5] Superviseurs (Directeurs) ──────────────────────")
            sup_map = await import_superviseurs(session, rows, school_map)

            print("── [5/5] Liaisons superviseurs ↔ enseignants ────────────")
            await link_superviseurs_to_teachers(session, school_map, sup_map)

            await session.commit()
            print("\n✅  Import terminé avec succès !")
            print(f"   Mot de passe temporaire des superviseurs : {SUPERVISOR_DEFAULT_PASSWORD}")
            print("   (les superviseurs devront le changer à la première connexion)\n")

        except Exception as e:
            await session.rollback()
            print(f"\n❌  Erreur — rollback effectué : {e}")
            import traceback
            traceback.print_exc()
            raise


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(
            "Usage :\n"
            "  Local  : .venv/bin/python scripts/import_ndawwune.py <fichier.xlsx> [--dry-run]\n"
            "  Docker : docker cp NdawWune.xlsx backend:/tmp/NdawWune.xlsx\n"
            "           docker exec backend python scripts/import_ndawwune.py /tmp/NdawWune.xlsx"
        )
        sys.exit(1)

    xlsx = sys.argv[1]
    dry = "--dry-run" in sys.argv

    if not Path(xlsx).exists():
        print(f"❌ Fichier introuvable : {xlsx}")
        sys.exit(1)

    asyncio.run(run(xlsx, dry_run=dry))
