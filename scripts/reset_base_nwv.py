#!/usr/bin/env python3
"""
Réinitialisation + réimport complet — BaseNWVFinale2026VL.xlsx
==============================================================
Remet la base au périmètre exact du fichier de référence : 6 IEF, 77 écoles,
79 tuteurs, 77 superviseurs, 2 966 élèves.

⚠️  DESTRUCTIF. Irréversible sans le dump que le script prend avant d'agir.

Ce qui est SUPPRIMÉ
    Structure  : schools, school_classes, eleves
    Terrain    : rapports_journalier (+ résolutions de difficultés), seances
                 (+ rapports_prof), supervisor_presence_checks,
                 evaluations_eleves, evaluation_tirages (+ evaluation_resultats),
                 eleve_remplacements, remarques, teacher_sessions, usage_logs
    Config      : progression_configs rattachées à une école — le repli global
                 (NULL, NULL) est PRÉSERVÉ (cf. CLAUDE.md §2 : sans lui, la
                 config par défaut du dashboard est silencieusement ignorée)
    Comptes    : tuteurs et superviseurs ABSENTS du fichier.
                 Les rôles admin et coordonnateur ne sont JAMAIS touchés.

Ce qui est PRÉSERVÉ
    Les comptes du fichier, avec téléphone, mot de passe, groupe_recherche et
    app_access intacts : personne n'est déconnecté. Seuls school_id, classes et
    niveau sont réécrits d'après le fichier.
    program_sessions, planning_segments, rapport_questions, rapport_libelles,
    rapport_difficultes, evaluation_docs, evaluation_sujets,
    evaluation_competences, documents, audit_logs.

Correspondances
    École       : code_ecole            Élève : code_eleve
    Classe      : (école, classe_eleve) — la classe RÉELLE de l'élève, plus fine
                  que la classe NDAW WUNE (CE1 A / CE1 B / CP…)
    Tuteur / superviseur : par nom, insensible aux accents, à la casse et aux
                  espaces multiples (le fichier écrit « SATOU SENE » là où le
                  dashboard a saisi « SATOU SÉNE »).

Usage — depuis le dossier backend/
    python3 scripts/reset_base_nwv.py ~/Downloads/BaseNWVFinale2026VL.xlsx
        Simulation : n'écrit rien, affiche tout ce qui serait fait.

    python3 scripts/reset_base_nwv.py ~/Downloads/BaseNWVFinale2026VL.xlsx \\
        --resoudre-doublons --creer-comptes-manquants \\
        --apply --confirm-purge OUI-SUPPRIMER-TOUT --confirm-db ared_ndawune
        Application réelle. La simulation affiche la commande exacte à recopier.

Options
    --apply                    Applique (sinon simulation + rollback)
    --confirm-purge TOKEN      Exigé avec --apply : OUI-SUPPRIMER-TOUT
    --confirm-db NOM           Exigé avec --apply : nom exact de la base visée.
                               Une DATABASE_URL mal transmise retombe sinon en
                               silence sur la base par défaut — c'est arrivé.
    --resoudre-doublons        Quand plusieurs comptes portent le même nom,
                               retient celui dont l'école correspond au fichier
                               et qui a un téléphone. Sans cette option, le
                               script s'arrête et affiche les candidats.
    --garder-compte UUID       Force la fiche à conserver quand le départage
                               automatique est impossible (deux fiches également
                               plausibles). Répétable. L'autre fiche est
                               supprimée comme compte hors fichier.
    --creer-comptes-manquants  Crée SANS TÉLÉPHONE les acteurs du fichier qui
                               n'ont pas de compte. ⚠️  Un compte sans téléphone
                               NE PEUT PAS SE CONNECTER (le téléphone est
                               l'identifiant) : il faudra le saisir ensuite dans
                               le dashboard. Sans cette option, le script
                               s'arrête et liste les manquants.
    --conserver-comptes-hors-fichier
                               Garde les comptes absents du fichier (école et
                               classes vidées) au lieu de les supprimer.
    --dump-dir DIR             Répertoire du dump (défaut : ../backups)
    --no-dump                  Saute le dump — seulement si un dump vient d'être pris
    --dump-container NOM       Conteneur Postgres si le pg_dump de l'hôte est
                               d'une version antérieure (défaut : backend-db-1)
"""

import argparse
import os
import re
import secrets
import subprocess
import sys
import unicodedata
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Forcer DATABASE_URL vers localhost hors Docker (même trick que import_eleves_2026.py) ──
_DEFAULT_DB = "postgresql+asyncpg://ared_user:ared_secret@localhost:5432/ared_ndawune"
if not os.path.exists("/.dockerenv") and (
    "DATABASE_URL" not in os.environ
    or "://db:" in os.environ.get("DATABASE_URL", "")
):
    os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL", _DEFAULT_DB).replace(
        "@db:", "@localhost:"
    )
    if os.environ["DATABASE_URL"].startswith("postgresql+asyncpg://$"):
        os.environ["DATABASE_URL"] = _DEFAULT_DB

import asyncio

from sqlalchemy import text

from app.core.database import AsyncSessionLocal
from app.core.security import hash_password


CONFIRM_TOKEN = "OUI-SUPPRIMER-TOUT"

# Ordre de suppression : des feuilles vers les racines. Les CASCADE feraient le
# travail pour l'essentiel, mais on supprime explicitement pour pouvoir compter
# ce qui part et ne rien détruire par effet de bord non prévu.
PURGE_ORDER = [
    ("rapport_difficulte_resolutions", None),
    ("rapports_journalier",            None),
    ("rapports_prof",                  None),
    ("seances",                        None),
    ("supervisor_presence_checks",     None),
    ("evaluations_eleves",             None),
    ("evaluation_resultats",           None),
    ("evaluation_tirages",             None),
    ("eleve_remplacements",            None),
    ("remarques",                      None),
    ("teacher_sessions",               None),
    ("usage_logs",                     None),
    ("progression_configs",            "school_id IS NOT NULL"),
    ("eleves",                         None),
    ("school_classes",                 None),
    ("schools",                        None),
]

REQUIRED_COLS = [
    "code_ecole", "nom_ecole", "nom_ief", "nom_commune",
    "code_superviseur", "nom_superviseur",
    "code_tuteur", "nom_tuteur",
    "classe_eleve", "niveau",
    "code_eleve", "nom_eleve", "sexe_str",
    "statut_selection", "groupe_lecture", "groupe_maths",
]


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def split_name(full_name):
    """(prénom, nom) — dernier mot = nom. Convention de import_eleves.py."""
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return ("", parts[0])
    return (" ".join(parts[:-1]), parts[-1])


def normalize_genre(sexe_str):
    s = str(sexe_str).strip().lower()
    if s in ("fille", "f"):
        return "Fille"
    if s in ("garcon", "garçon", "g", "m"):
        return "Garçon"
    return str(sexe_str).strip()


def norm_name(n):
    """Clé de correspondance d'un nom de personne : insensible à la casse, aux
    accents et aux espaces multiples.

    Le fichier écrit les noms sans accent (SATOU SENE) là où le dashboard les a
    saisis avec (SATOU SÉNE), et certains comptes portent un double espace
    (« MODOU  FALL »). Une comparaison stricte déclarait 24 acteurs sur 25
    « absents de la base » alors qu'ils existaient bel et bien.
    """
    if n is None:
        return ""
    s = unicodedata.normalize("NFD", str(n))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return " ".join(s.split()).upper()


def code_str(v):
    if v is None or str(v).strip() == "":
        return ""
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v).strip()


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ──────────────────────────────────────────────────────────────────────────────
# Lecture du fichier
# ──────────────────────────────────────────────────────────────────────────────

def load_source(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {name: header.index(name) for name in REQUIRED_COLS if name in header}
    manquantes = [c for c in REQUIRED_COLS if c not in idx]
    if manquantes:
        sys.exit(f"❌  Colonnes manquantes dans le fichier : {manquantes}")
    i_langue = header.index("Langue") if "Langue" in header else None

    def g(row, name):
        i = idx[name]
        return row[i] if i < len(row) else None

    ecoles, classes, superviseurs, tuteurs, lien, eleves = {}, {}, {}, {}, {}, []

    for row in rows:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        c_ecole = code_str(g(row, "code_ecole"))
        c_sup   = code_str(g(row, "code_superviseur"))
        c_tut   = code_str(g(row, "code_tuteur"))
        c_elv   = code_str(g(row, "code_eleve"))
        classe  = txt(g(row, "classe_eleve"))
        niveau  = txt(g(row, "niveau"))

        if c_ecole and c_ecole not in ecoles:
            langue = row[i_langue] if i_langue is not None and i_langue < len(row) else None
            ecoles[c_ecole] = {
                "name":   txt(g(row, "nom_ecole")),
                "region": txt(g(row, "nom_ief")),
                "city":   txt(g(row, "nom_commune")),
                "langue": (txt(langue) or "").lower() or None,
            }

        if c_ecole and classe:
            key = (c_ecole, classe)
            if key not in classes:
                classes[key] = {"niveau": niveau, "effectif": 0}
            classes[key]["effectif"] += 1

        if c_sup and c_sup not in superviseurs:
            superviseurs[c_sup] = {"name": txt(g(row, "nom_superviseur")), "code_ecole": c_ecole}

        if c_tut:
            if c_tut not in tuteurs:
                tuteurs[c_tut] = {"name": txt(g(row, "nom_tuteur")), "code_ecole": c_ecole,
                                  "classes": set(), "niveau": set()}
            if classe:
                tuteurs[c_tut]["classes"].add(classe)
            if niveau:
                tuteurs[c_tut]["niveau"].add(niveau)

        if c_sup and c_tut:
            lien.setdefault(c_sup, set()).add(c_tut)

        if c_elv:
            prenom, nom = split_name(txt(g(row, "nom_eleve")))
            eleves.append({
                "code_eleve": c_elv, "nom": nom, "prenom": prenom or None,
                "classe": classe, "code_ecole": c_ecole,
                "genre": normalize_genre(g(row, "sexe_str")),
                "statut_selection": txt(g(row, "statut_selection")),
                "groupe_lecture":   txt(g(row, "groupe_lecture")),
                "groupe_maths":     txt(g(row, "groupe_maths")),
            })
    wb.close()

    # Homonymes : uq_eleve_school_classe_nom porte sur (school_id, classe, nom,
    # prenom). Deux élèves distincts (codes différents) du même nom dans la même
    # classe violeraient la contrainte — l'un des deux serait rejeté EN SILENCE,
    # ce qui s'est déjà produit sur 8 élèves. On désambiguïse par le code_eleve
    # dans le prénom, ce qui les distingue aussi pour le tuteur dans l'app.
    vus, homonymes = {}, []
    for e in eleves:
        vus.setdefault((e["code_ecole"], e["classe"], e["nom"], e["prenom"]), []).append(e)
    for groupe in vus.values():
        if len(groupe) > 1:
            for e in groupe:
                e["prenom"] = f"{e['prenom'] or ''} ({e['code_eleve']})".strip()
                homonymes.append(e)

    return {"ecoles": ecoles, "classes": classes, "superviseurs": superviseurs,
            "tuteurs": tuteurs, "lien": lien, "eleves": eleves, "homonymes": homonymes}


# ──────────────────────────────────────────────────────────────────────────────
# Dump préalable
# ──────────────────────────────────────────────────────────────────────────────

def db_params():
    """Extrait host/port/db/user/password de DATABASE_URL."""
    m = re.match(r".*://([^:]+):([^@]*)@([^:/]+):(\d+)/(.+)$", os.environ["DATABASE_URL"])
    if not m:
        return None
    user, password, host, port, dbname = m.groups()
    return {"user": user, "password": password, "host": host,
            "port": port, "dbname": dbname.split("?")[0]}


def take_dump(dump_dir, container):
    p = db_params()
    if p is None:
        sys.exit("❌  DATABASE_URL illisible — impossible de prendre un dump.")
    d = Path(dump_dir)
    d.mkdir(parents=True, exist_ok=True)
    out = d / f"avant-reset-{datetime.now():%Y%m%d-%H%M%S}.dump"
    print(f"\n💾  Dump préalable → {out}")

    env = {"PGPASSWORD": p["password"], "PATH": os.environ.get("PATH", "/usr/bin:/bin")}
    ok, err = False, ""
    try:
        r = subprocess.run(
            ["pg_dump", "-Fc", "-h", p["host"], "-p", p["port"], "-U", p["user"],
             "-d", p["dbname"], "-f", str(out)],
            env=env, capture_output=True, text=True)
        ok, err = r.returncode == 0, (r.stderr or "").strip()
    except FileNotFoundError:
        err = "pg_dump absent de l'hôte"

    if not ok:
        # Le Postgres tourne en conteneur et le pg_dump de l'hôte est souvent
        # d'une version antérieure — pg_dump refuse alors de tourner.
        print(f"   pg_dump hôte indisponible ({err.splitlines()[-1] if err else '?'})"
              f" — repli sur le conteneur « {container} »…")
        try:
            with out.open("wb") as fh:
                r = subprocess.run(
                    ["docker", "exec", "-e", f"PGPASSWORD={p['password']}", container,
                     "pg_dump", "-Fc", "-U", p["user"], "-d", p["dbname"]],
                    stdout=fh, stderr=subprocess.PIPE)
            ok, err = r.returncode == 0, (r.stderr or b"").decode().strip()
        except FileNotFoundError:
            ok, err = False, "docker absent de l'hôte"

    if not ok:
        out.unlink(missing_ok=True)
        # Cas courant sur le VPS : le script tourne dans le conteneur backend,
        # dont l'image n'embarque ni pg_dump (Dockerfile : libpq-dev seulement)
        # ni la CLI docker. Le dump doit alors être pris depuis l'hôte.
        dans_conteneur = os.path.exists("/.dockerenv")
        msg = [f"❌  Dump impossible : {err}", "   RIEN n'a été supprimé."]
        if dans_conteneur:
            msg += [
                "",
                "   Ce script tourne dans un conteneur, qui n'a pas pg_dump.",
                "   Prenez le dump depuis l'HÔTE, puis relancez avec --no-dump :",
                "",
                "     cd ~/ndaw-wune && set -a && . ./.env && set +a && mkdir -p backups \\",
                "       && docker compose exec -T -e PGPASSWORD=\"$POSTGRES_PASSWORD\" db \\",
                f"          pg_dump -U \"$POSTGRES_USER\" -d \"$POSTGRES_DB\" -Fc \\",
                "          > backups/avant-reset-$(date +%Y%m%d-%H%M%S).dump",
            ]
        else:
            msg.append("   Prenez un dump manuellement puis relancez avec --no-dump.")
        sys.exit("\n".join(msg))
    taille = out.stat().st_size
    if taille < 1024:
        out.unlink(missing_ok=True)
        sys.exit(f"❌  Dump suspect ({taille} octets) — rien n'a été supprimé.")
    print(f"   OK — {taille / 1024 / 1024:.1f} Mo")
    return out


# ──────────────────────────────────────────────────────────────────────────────
# Traitement
# ──────────────────────────────────────────────────────────────────────────────

async def run(args):
    tag = "" if args.apply else "[SIMULATION] "
    print(f"\n{tag}Réinitialisation + réimport — {args.fichier}")
    print("=" * 72)

    print("📂  Lecture du fichier Excel…")
    src = load_source(args.fichier)
    print(f"   {len(src['ecoles'])} écoles · {len(src['classes'])} classes · "
          f"{len(src['superviseurs'])} superviseurs · {len(src['tuteurs'])} tuteurs · "
          f"{len(src['eleves'])} élèves")
    if src["homonymes"]:
        print(f"   ℹ️  {len(src['homonymes'])} élèves homonymes dans une même classe — "
              f"prénom suffixé du code_eleve pour les distinguer.")

    p = db_params()
    print(f"\n🔌  Connexion à {p['dbname']}@{p['host']}:{p['port']}…" if p else "\n🔌  Connexion…")

    async with AsyncSessionLocal() as s:
        # ── 1. Correspondance des comptes ─────────────────────────────────────
        print("\n🔎  Contrôle des comptes tuteurs / superviseurs…")
        rows = (await s.execute(text(
            "SELECT u.id, u.name, u.phone, u.role, u.created_at, s.name AS ecole, u.classes "
            "FROM users u LEFT JOIN schools s ON s.id = u.school_id "
            "WHERE u.role IN ('enseignant', 'superviseur')"
        ))).mappings().all()

        par_nom = {"enseignant": {}, "superviseur": {}}
        for r in rows:
            par_nom[r["role"]].setdefault(norm_name(r["name"]), []).append(r)

        ecole_attendue = {}
        for role, coll in (("enseignant", src["tuteurs"]), ("superviseur", src["superviseurs"])):
            for info in coll.values():
                ec = src["ecoles"].get(info["code_ecole"], {})
                ecole_attendue[(role, norm_name(info["name"]))] = ec.get("name", "?")

        forces = {u.lower() for u in args.garder_compte}

        def choisir(role, key, candidats):
            """Retient le compte le plus plausible : école conforme au fichier
            (+2), téléphone renseigné (+1), à égalité le plus ancien.

            --garder-compte <uuid> force le choix et court-circuite le score :
            c'est la seule issue quand deux fiches sont également plausibles
            (même école, deux téléphones valides), car seul un humain sait
            lequel des deux numéros la personne utilise réellement.
            """
            attendue = ecole_attendue.get((role, key))
            impose = [r for r in candidats if str(r["id"]).lower() in forces]
            if len(impose) == 1:
                return impose[0], attendue
            def score(r):
                return ((2 if r["ecole"] == attendue else 0)
                        + (1 if r["phone"] else 0))
            classe = sorted(candidats, key=lambda r: (-score(r), r["created_at"]))
            if len(classe) > 1 and score(classe[0]) == score(classe[1]):
                return None, attendue   # départage impossible
            return classe[0], attendue

        retenus = {"enseignant": {}, "superviseur": {}}
        ambigus, non_retenus = [], []
        for role in ("enseignant", "superviseur"):
            attendus = {norm_name(i["name"])
                        for i in (src["tuteurs"] if role == "enseignant"
                                  else src["superviseurs"]).values()}
            for key, candidats in par_nom[role].items():
                if key not in attendus:
                    continue
                if len(candidats) == 1:
                    retenus[role][key] = candidats[0]
                    continue
                gagnant, attendue = choisir(role, key, candidats)
                if gagnant is None or not args.resoudre_doublons:
                    ambigus.append((role, key, candidats, attendue, gagnant))
                else:
                    retenus[role][key] = gagnant
                    non_retenus += [c for c in candidats if c["id"] != gagnant["id"]]

        if ambigus:
            print("\n❌  COMPTES AMBIGUS — plusieurs comptes portent le même nom une fois "
                  "accents et espaces neutralisés :")
            indecidables = []
            for role, key, candidats, attendue, gagnant in ambigus:
                print(f"\n      [{role}] {key} — le fichier l'attend à « {attendue} »")
                for c in candidats:
                    tel = c["phone"] or "SANS TÉLÉPHONE (ne peut pas se connecter)"
                    marque = "  ← correspond" if c["ecole"] == attendue else ""
                    retenu = "  ← SERAIT RETENU" if gagnant and c["id"] == gagnant["id"] else ""
                    print(f"        id={c['id']}")
                    print(f"          « {c['name']} » · {tel} · {c['ecole']}{marque}{retenu}")
                if gagnant is None:
                    indecidables.append((role, key))
                    print("        → départage automatique impossible : les deux fiches sont")
                    print("          également plausibles. Seul un humain sait lequel des")
                    print("          numéros la personne utilise pour se connecter.")
            if indecidables and args.resoudre_doublons:
                print("\n   → Choisissez explicitement la fiche à garder :")
                print("        --garder-compte <id>        (répétable)")
                print("     L'autre sera supprimée comme compte hors fichier.")
            else:
                print("\n   → Relancez avec --resoudre-doublons pour retenir automatiquement")
                print("     le compte marqué, ou corrigez les fiches dans le dashboard.")
            return

        # ── 2. Acteurs du fichier sans compte ─────────────────────────────────
        a_creer = []
        for role, coll in (("enseignant", src["tuteurs"]), ("superviseur", src["superviseurs"])):
            for info in coll.values():
                if norm_name(info["name"]) not in retenus[role]:
                    a_creer.append((role, info["name"]))

        print(f"   Tuteurs      : {len(retenus['enseignant'])}/{len(src['tuteurs'])} "
              f"retrouvés en base")
        print(f"   Superviseurs : {len(retenus['superviseur'])}/{len(src['superviseurs'])} "
              f"retrouvés en base")

        if a_creer and not args.creer_comptes_manquants:
            print("\n❌  ACTEURS DU FICHIER SANS COMPTE EN BASE")
            print("   Le fichier ne contient aucun téléphone de tuteur ni de superviseur.")
            print("   Un compte créé sans téléphone NE PEUT PAS SE CONNECTER : le téléphone")
            print("   est l'identifiant. La purge laisserait ces écoles sans personne pour")
            print("   saisir, jusqu'à ce qu'un numéro leur soit ajouté dans le dashboard.")
            for role, nom in a_creer:
                print(f"      [{role}] {nom}")
            print("\n   → Créez-les dans le dashboard avec leur vrai numéro, corrigez")
            print("     l'orthographe dans le fichier, ou relancez avec")
            print("     --creer-comptes-manquants pour les créer sans téléphone.")
            return

        if a_creer:
            print(f"\n   ⚠️  {len(a_creer)} compte(s) seront CRÉÉS SANS TÉLÉPHONE "
                  f"→ connexion impossible tant qu'un numéro n'est pas saisi :")
            for role, nom in a_creer:
                print(f"        [{role}] {nom}")

        # ── 3. Comptes hors fichier ───────────────────────────────────────────
        orphelins = list(non_retenus)
        gardes = {str(r["id"]) for role in retenus for r in retenus[role].values()}
        for r in rows:
            if str(r["id"]) not in gardes and not any(str(r["id"]) == str(o["id"])
                                                      for o in orphelins):
                orphelins.append(r)

        if orphelins:
            verbe = "CONSERVÉS" if args.conserver_comptes_hors_fichier else "SUPPRIMÉS"
            n_tut = sum(1 for r in orphelins if r["role"] == "enseignant")
            n_sup = len(orphelins) - n_tut
            print(f"\n   {n_tut} tuteur(s) et {n_sup} superviseur(s) en base sont "
                  f"ABSENTS du fichier → {verbe}")
            for r in sorted(orphelins, key=lambda x: (x["role"], x["name"])):
                print(f"        [{r['role']}] {r['name']}")

            if not args.conserver_comptes_hors_fichier:
                ids = [str(r["id"]) for r in orphelins]
                print("\n      Effets de bord (FK ON DELETE SET NULL) — lignes conservées,")
                print("      elles perdent seulement leur auteur :")
                for table, col in (("planning_segments", "teacher_id"),
                                   ("documents", "uploaded_by"),
                                   ("evaluation_sujets", "created_by")):
                    n = (await s.execute(text(
                        f"SELECT COUNT(*) FROM {table} WHERE {col}::text = ANY(:ids)"
                    ), {"ids": ids})).scalar_one()
                    if n:
                        print(f"        {table}.{col} → NULL sur {n} ligne(s)")

        # ── 4. Inventaire ─────────────────────────────────────────────────────
        print("\n🗑️   Ce qui va être SUPPRIMÉ :")
        total = 0
        for table, where in PURGE_ORDER:
            clause = f" WHERE {where}" if where else ""
            n = (await s.execute(text(f"SELECT COUNT(*) FROM {table}{clause}"))).scalar_one()
            total += n
            print(f"   {table:35s} {n:7d}{'' if n == 0 else '  ←'}")
        if orphelins and not args.conserver_comptes_hors_fichier:
            print(f"   {'users (enseignant/superviseur)':35s} {len(orphelins):7d}  ←")
            total += len(orphelins)
        print(f"   {'TOTAL':35s} {total:7d} lignes")

        restants = {r["role"]: r["n"] for r in (await s.execute(text(
            "SELECT role, COUNT(*) AS n FROM users GROUP BY role"))).mappings().all()}
        if not args.conserver_comptes_hors_fichier:
            for r in orphelins:
                restants[r["role"]] = restants.get(r["role"], 0) - 1
        for role, nom in a_creer:
            restants[role] = restants.get(role, 0) + 1
        print("\n   Comptes après l'opération :")
        for role in sorted(restants):
            print(f"      {role:15s} {restants[role]:5d}")

        if not args.apply:
            print("\n" + "=" * 72)
            print("🧪  SIMULATION — rien n'a été écrit.")
            print("    Pour appliquer, ajoutez :")
            print(f"      --apply --confirm-purge {CONFIRM_TOKEN} --confirm-db {p['dbname']}")
            await s.rollback()
            return

        # ── Garde-fou de cible ────────────────────────────────────────────────
        # Une variable DATABASE_URL mal transmise (shell, .env non chargé, faute
        # de frappe) fait silencieusement retomber le script sur la base par
        # défaut. Exiger le nom de la base en clair rend la cible impossible à
        # confondre : le token de purge seul ne dit pas SUR QUOI on purge.
        if args.confirm_db != p["dbname"]:
            print(f"\n❌  Cible non confirmée. Le script est connecté à « {p['dbname']} ».")
            print(f"    Ajoutez --confirm-db {p['dbname']} si c'est bien la base à purger.")
            await s.rollback()
            sys.exit(1)
        print(f"\n🎯  Cible confirmée : {p['dbname']}@{p['host']}")

        # ── 5. Dump ───────────────────────────────────────────────────────────
        if args.no_dump:
            print("\n⚠️   --no-dump : aucun dump pris par ce script.")
        else:
            take_dump(args.dump_dir, args.dump_container)

        # ── 6. Purge ──────────────────────────────────────────────────────────
        print("\n🗑️   Purge…")
        for table, where in PURGE_ORDER:
            clause = f" WHERE {where}" if where else ""
            r = await s.execute(text(f"DELETE FROM {table}{clause}"))
            if r.rowcount:
                print(f"   {table:35s} -{r.rowcount}")

        # ── 7. Session active ─────────────────────────────────────────────────
        row = (await s.execute(text(
            "SELECT id, name FROM program_sessions WHERE status = 'active' "
            "ORDER BY created_at DESC LIMIT 1"))).mappings().first()
        session_id = str(row["id"]) if row else None
        print(f"\n📅  Session active : « {row['name']} »" if row
              else "\n   ⚠️  Aucune session active — élèves créés sans session_id.")

        # ── 8. Écoles ─────────────────────────────────────────────────────────
        print("\n🏫  Écoles…")
        school_id_by_code = {}
        for code, info in src["ecoles"].items():
            sid = str(uuid.uuid4())
            await s.execute(text(
                "INSERT INTO schools (id, name, code_ecole, region, city, langue, "
                "created_at, updated_at) VALUES (:id, :name, :code, :region, :city, "
                ":langue, NOW(), NOW())"),
                {"id": sid, "name": info["name"], "code": int(code),
                 "region": info["region"], "city": info["city"], "langue": info["langue"]})
            school_id_by_code[code] = sid
        print(f"   {len(school_id_by_code)} créée(s).")

        # ── 9. Classes ────────────────────────────────────────────────────────
        print("📚  Classes…")
        n_cl = 0
        for (code_ecole, nom_classe), info in src["classes"].items():
            sid = school_id_by_code.get(code_ecole)
            if sid is None:
                continue
            await s.execute(text(
                "INSERT INTO school_classes (id, school_id, name, niveau, effectif, "
                "created_at, updated_at) VALUES (:id, :sid, :name, :niveau, :eff, NOW(), NOW())"),
                {"id": str(uuid.uuid4()), "sid": sid, "name": nom_classe,
                 "niveau": info["niveau"], "eff": info["effectif"]})
            n_cl += 1
        print(f"   {n_cl} créée(s).")

        # ── 10. Comptes hors fichier ──────────────────────────────────────────
        if orphelins and not args.conserver_comptes_hors_fichier:
            # Restreint aux rôles enseignant/superviseur : un compte admin ou
            # coordonnateur n'apparaît pas dans le fichier et ne doit pas partir.
            r = await s.execute(text(
                "DELETE FROM users WHERE id::text = ANY(:ids) "
                "AND role IN ('enseignant', 'superviseur')"),
                {"ids": [str(o["id"]) for o in orphelins]})
            print(f"👤  Comptes hors fichier supprimés : {r.rowcount}")

        # ── 11. Comptes manquants ─────────────────────────────────────────────
        for role, nom in a_creer:
            uid = str(uuid.uuid4())
            await s.execute(text(
                "INSERT INTO users (id, name, phone, password_hash, role, status, "
                "app_access, must_change_password, created_at, updated_at) "
                "VALUES (:id, :name, NULL, :h, :role, 'actif', 'full', true, NOW(), NOW())"),
                {"id": uid, "name": nom, "h": hash_password(secrets.token_urlsafe(16)),
                 "role": role})
            retenus[role][norm_name(nom)] = {"id": uid, "name": nom}
        if a_creer:
            print(f"👤  Comptes créés sans téléphone : {len(a_creer)}")

        # ── 12. Réaffectation ─────────────────────────────────────────────────
        print("👤  Réaffectation des comptes (école, classes, niveau)…")
        await s.execute(text(
            "UPDATE users SET school_id = NULL, classes = NULL, niveau = NULL, "
            "updated_at = NOW() WHERE role IN ('enseignant', 'superviseur')"))

        tut_id_by_code = {}
        for code, info in src["tuteurs"].items():
            u = retenus["enseignant"][norm_name(info["name"])]
            tut_id_by_code[code] = str(u["id"])
            await s.execute(text(
                "UPDATE users SET school_id = :sid, classes = :cl, niveau = :ni, "
                "updated_at = NOW() WHERE id = :id"),
                {"sid": school_id_by_code.get(info["code_ecole"]),
                 "cl": sorted(info["classes"]), "ni": sorted(info["niveau"]),
                 "id": str(u["id"])})

        sup_id_by_code = {}
        for code, info in src["superviseurs"].items():
            u = retenus["superviseur"][norm_name(info["name"])]
            sup_id_by_code[code] = str(u["id"])
            await s.execute(text(
                "UPDATE users SET school_id = :sid, updated_at = NOW() WHERE id = :id"),
                {"sid": school_id_by_code.get(info["code_ecole"]), "id": str(u["id"])})
        print(f"   {len(tut_id_by_code)} tuteur(s), {len(sup_id_by_code)} superviseur(s).")

        # ── 13. Lien superviseur → enseignants ────────────────────────────────
        # users.classes d'un SUPERVISEUR contient des UUID d'enseignants, pas des
        # noms de classes (cf. CLAUDE.md §2).
        print("🔗  Rattachement superviseur → enseignants…")
        n_liens = 0
        for code_sup, codes_tut in src["lien"].items():
            sup_id = sup_id_by_code.get(code_sup)
            ids = sorted(tut_id_by_code[c] for c in codes_tut if c in tut_id_by_code)
            if not sup_id or not ids:
                continue
            await s.execute(text(
                "UPDATE users SET classes = :ids, updated_at = NOW() WHERE id = :id"),
                {"ids": ids, "id": sup_id})
            n_liens += 1
        print(f"   {n_liens} superviseur(s) relié(s).")

        # ── 14. Élèves ────────────────────────────────────────────────────────
        print("🧒  Élèves…")
        lots, ignores = [], 0
        for e in src["eleves"]:
            sid = school_id_by_code.get(e["code_ecole"])
            if sid is None:
                ignores += 1
                continue
            lots.append({"id": str(uuid.uuid4()), "nom": e["nom"], "prenom": e["prenom"],
                         "code": e["code_eleve"], "classe": e["classe"], "genre": e["genre"],
                         "sel": e["statut_selection"], "lec": e["groupe_lecture"],
                         "mat": e["groupe_maths"], "sid": sid, "sess": session_id})
        await s.execute(text(
            "INSERT INTO eleves (id, nom, prenom, code_eleve, classe, genre, statut, "
            "statut_selection, groupe_lecture, groupe_maths, school_id, session_id, "
            "created_at, updated_at) VALUES (:id, :nom, :prenom, :code, :classe, :genre, "
            "'actif', :sel, :lec, :mat, :sid, :sess, NOW(), NOW())"), lots)
        print(f"   {len(lots)} créé(s)" + (f", {ignores} ignoré(s)." if ignores else "."))

        # ── 15. Contrôles post-import ─────────────────────────────────────────
        print("\n✅  Contrôles post-import…")
        orphelins_eleves = (await s.execute(text("""
            SELECT COUNT(*) FROM eleves e WHERE NOT EXISTS (
                SELECT 1 FROM users u WHERE u.role = 'enseignant'
                  AND u.school_id = e.school_id AND e.classe = ANY(u.classes))
        """))).scalar_one()
        if orphelins_eleves:
            print(f"   ❌  {orphelins_eleves} élève(s) rattaché(s) à AUCUN tuteur — "
                  f"ils seraient invisibles dans l'app. Annulation.")
            await s.rollback()
            sys.exit(1)
        print("   Tous les élèves sont rattachés à un tuteur ✓")

        n = (await s.execute(text(
            "SELECT COUNT(*) FROM users WHERE role = 'superviseur' "
            "AND (classes IS NULL OR cardinality(classes) = 0)"))).scalar_one()
        print(f"   Superviseurs sans enseignant assigné : {n}")

        await s.commit()
        print("\n" + "=" * 72)
        print("✅  Appliqué et validé (commit).")
        if a_creer:
            print(f"   ⚠️  {len(a_creer)} compte(s) sans téléphone — saisissez leur numéro")
            print("      dans le dashboard, sinon ils ne pourront pas se connecter.")
        print("   Vérifiez le dashboard, puis qu'un tuteur voit bien ses élèves dans l'app.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("fichier", help="Chemin vers BaseNWVFinale2026VL.xlsx")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--confirm-purge", default="")
    ap.add_argument("--confirm-db", default="", metavar="NOM",
                    help="Nom exact de la base à purger. Exigé avec --apply : "
                         "protège contre une DATABASE_URL mal transmise.")
    ap.add_argument("--resoudre-doublons", action="store_true")
    ap.add_argument("--garder-compte", action="append", default=[], metavar="UUID",
                    help="Force la fiche à conserver quand plusieurs comptes portent "
                         "le même nom. Répétable.")
    ap.add_argument("--creer-comptes-manquants", action="store_true")
    ap.add_argument("--conserver-comptes-hors-fichier", action="store_true")
    ap.add_argument("--dump-dir", default=str(ROOT.parent / "backups"))
    ap.add_argument("--no-dump", action="store_true")
    ap.add_argument("--dump-container", default="backend-db-1")
    args = ap.parse_args()

    if args.apply and args.confirm_purge != CONFIRM_TOKEN:
        sys.exit(f"❌  --apply exige --confirm-purge {CONFIRM_TOKEN}")

    asyncio.run(run(args))
