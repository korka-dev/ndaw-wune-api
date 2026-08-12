#!/usr/bin/env python3
"""Vérification du round-trip Import/Export du dashboard admin.

    exporter → ajouter des lignes au fichier → réimporter

Contrôle la promesse de la fonctionnalité : les nouvelles données sont ajoutées,
les existantes ignorées, le réimport d'un export intact ne crée rien, et surtout
les élèves ajoutés sont VISIBLES par leur tuteur — un élève créé dans une classe
absente de users.classes existe en base mais n'apparaît jamais dans l'app
(cf. CLAUDE.md §2).

⚠️  Le script ÉCRIT en base. À lancer sur une copie, jamais sur la production :

    createdb roundtrip_test
    pg_dump -d ared_ndawune | psql -d roundtrip_test
    DATABASE_URL=postgresql+asyncpg://user:pass@localhost:5432/roundtrip_test \
        python3 scripts/test_import_export_roundtrip.py
"""
import asyncio, io, os, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

if "roundtrip" not in os.environ.get("DATABASE_URL", ""):
    sys.exit("❌  Refus : DATABASE_URL doit viser une base de test dont le nom "
             "contient « roundtrip ». Ce script écrit en base.")

import openpyxl
from fastapi import UploadFile
from sqlalchemy import text

from app.core.database import AsyncSessionLocal
from app.api.routes.admin.import_export import export_all, import_all, COLUMNS


async def exporter(db):
    resp = await export_all(db, None)
    buf = b"".join([c async for c in resp.body_iterator] if hasattr(resp.body_iterator, "__aiter__")
                   else list(resp.body_iterator))
    return buf


def ajouter_lignes(contenu, lignes):
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb.active
    for l in lignes:
        ws.append([l.get(c) for c in COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def importer(db, contenu):
    up = UploadFile(file=io.BytesIO(contenu), filename="reimport.xlsx")
    return await import_all(db, None, up)


async def etat(db, label):
    q = await db.execute(text("""
        SELECT (SELECT count(*) FROM schools) AS ecoles,
               (SELECT count(*) FROM school_classes) AS classes,
               (SELECT count(*) FROM eleves) AS eleves,
               (SELECT count(*) FROM users WHERE role='enseignant') AS tuteurs,
               (SELECT count(*) FROM users WHERE role='superviseur') AS superviseurs
    """))
    r = q.mappings().first()
    print(f"   {label:22s} écoles={r['ecoles']} classes={r['classes']} "
          f"élèves={r['eleves']} tuteurs={r['tuteurs']} sup={r['superviseurs']}")


async def eleves_invisibles(db):
    """Élèves qu'aucun tuteur ne peut voir — le piège (school_id, classe)."""
    q = await db.execute(text("""
        SELECT e.nom, e.prenom, e.classe, s.name AS ecole
        FROM eleves e LEFT JOIN schools s ON s.id = e.school_id
        WHERE NOT EXISTS (
            SELECT 1 FROM users u WHERE u.role='enseignant'
              AND u.school_id = e.school_id AND e.classe = ANY(u.classes))
    """))
    return q.mappings().all()


async def main():
    async with AsyncSessionLocal() as db:
        print("── État initial")
        await etat(db, "avant")

        contenu = await exporter(db)
        wb = openpyxl.load_workbook(io.BytesIO(contenu))
        print(f"   export : {wb.active.max_row - 1} lignes")

        # ── Les ajouts à tester ───────────────────────────────────────────
        nouvelles = [
            # 1. École entièrement nouvelle, tuteur sur DEUX classes
            {"IEF": "IEF Test", "Commune": "Com Test", "Code École": 9001,
             "École": "EE TEST NOUVELLE", "Langue": "wolof",
             "Superviseur": "SUP TEST", "Téléphone Superviseur": "770000901",
             "Tuteur": "TUTEUR TEST", "Téléphone Tuteur": "770000902",
             "Classe": "CE1 A", "Niveau": "CE1", "Code Élève": "9001101",
             "Nom": "DIOP", "Prénom": "Awa", "Sexe": "Fille", "Type": "Titulaire",
             "Groupe Lecture": "Mots/Phrases", "Groupe Maths": "Nombres", "Statut": "actif"},
            {"IEF": "IEF Test", "Commune": "Com Test", "Code École": 9001,
             "École": "EE TEST NOUVELLE", "Langue": "wolof",
             "Superviseur": "SUP TEST", "Téléphone Superviseur": "770000901",
             "Tuteur": "TUTEUR TEST", "Téléphone Tuteur": "770000902",
             "Classe": "CE1 B", "Niveau": "CE1", "Code Élève": "9001102",
             "Nom": "FALL", "Prénom": "Modou", "Sexe": "Garcon", "Type": "Titulaire",
             "Groupe Lecture": "Lettres/Syllabes", "Groupe Maths": "Opérations",
             "Statut": "actif"},
        ]

        # 2. Nouvel élève dans une classe EXISTANTE d'un tuteur existant
        r = (await db.execute(text("""
            SELECT s.name AS ecole, s.code_ecole, s.region, s.city, s.langue,
                   u.name AS tuteur, u.phone, e.classe
            FROM eleves e JOIN schools s ON s.id = e.school_id
            JOIN users u ON u.school_id = s.id AND u.role='enseignant'
                        AND e.classe = ANY(u.classes)
            LIMIT 1"""))).mappings().first()
        nouvelles.append({
            "IEF": r["region"], "Commune": r["city"], "Code École": r["code_ecole"],
            "École": r["ecole"], "Langue": r["langue"],
            "Tuteur": r["tuteur"], "Téléphone Tuteur": r["phone"],
            "Classe": r["classe"], "Niveau": "CE1", "Code Élève": "9999001",
            "Nom": "NOUVEAU", "Prénom": "Eleve", "Sexe": "Fille",
            "Type": "Remplaçant", "Groupe Lecture": "Mots/Phrases",
            "Groupe Maths": "Nombres", "Statut": "actif"})

        # 3. Classe NOUVELLE pour un tuteur EXISTANT (le cas piégeux)
        nouvelles.append({
            "IEF": r["region"], "Commune": r["city"], "Code École": r["code_ecole"],
            "École": r["ecole"], "Langue": r["langue"],
            "Tuteur": r["tuteur"], "Téléphone Tuteur": r["phone"],
            "Classe": "CM2 Z", "Niveau": "CM2", "Code Élève": "9999002",
            "Nom": "NOUVEAU", "Prénom": "Classe", "Sexe": "Garcon",
            "Type": "Titulaire", "Groupe Lecture": "Mots/Phrases",
            "Groupe Maths": "Nombres", "Statut": "actif"})

        print(f"\n── Réimport de l'export + {len(nouvelles)} lignes ajoutées")
        res = await importer(db, ajouter_lignes(contenu, nouvelles))
        await db.commit()

        print(f"   écoles={res.ecoles_creees} sup={res.superviseurs_crees} "
              f"tuteurs={res.tuteurs_crees} classes={res.classes_creees} "
              f"élèves={res.eleves_crees} ignorés={res.ignores}")
        for e in res.erreurs[:5]:
            print(f"   ERREUR {e}")

        await etat(db, "après")

        print("\n── Vérifications")
        att = {"9001101": "EE TEST NOUVELLE/CE1 A", "9001102": "EE TEST NOUVELLE/CE1 B",
               "9999001": "classe existante", "9999002": "classe nouvelle"}
        for code, quoi in att.items():
            n = (await db.execute(text(
                "SELECT count(*) FROM eleves WHERE code_eleve=:c"), {"c": code})).scalar_one()
            print(f"   élève {code} ({quoi}) : {'créé ✓' if n else 'ABSENT ✗'}")

        t = (await db.execute(text(
            "SELECT classes FROM users WHERE upper(name)='TUTEUR TEST'"))).scalar_one_or_none()
        ok = t and set(t) == {"CE1 A", "CE1 B"}
        print(f"   TUTEUR TEST.classes = {t}  {'✓' if ok else '✗ attendu [CE1 A, CE1 B]'}")

        t2 = (await db.execute(text(
            "SELECT classes FROM users WHERE name=:n AND role='enseignant'"),
            {"n": r["tuteur"]})).scalar_one_or_none()
        ok2 = t2 and "CM2 Z" in t2
        print(f"   {r['tuteur']}.classes contient CM2 Z : {'✓' if ok2 else '✗ ' + str(t2)}")

        inv = await eleves_invisibles(db)
        print(f"\n   Élèves invisibles par leur tuteur : {len(inv)}")
        for e in inv[:8]:
            print(f"      {e['nom']} {e['prenom'] or ''} · {e['ecole']} · {e['classe']}")

        print("\n── Second réimport à l'identique (doit tout ignorer)")
        res2 = await importer(db, ajouter_lignes(await exporter(db), []))
        await db.commit()
        print(f"   écoles={res2.ecoles_creees} sup={res2.superviseurs_crees} "
              f"tuteurs={res2.tuteurs_crees} classes={res2.classes_creees} "
              f"élèves={res2.eleves_crees}")
        cree = (res2.ecoles_creees + res2.superviseurs_crees + res2.tuteurs_crees
                + res2.classes_creees + res2.eleves_crees)
        print(f"   {'✓ rien créé' if cree == 0 else f'✗ {cree} créations en double'}")


asyncio.run(main())
